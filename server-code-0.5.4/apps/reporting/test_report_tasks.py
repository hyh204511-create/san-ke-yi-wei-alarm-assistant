import hashlib
import tempfile
from collections import Counter
from datetime import timedelta
from io import BytesIO
from zipfile import ZipFile

from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.utils import timezone
from openpyxl import load_workbook

from apps.governance.models import AssistantProfile, EnterpriseGrant, EnterpriseScope, RoleAssignment
from apps.governance.services import assign_role, claim_shift

from . import report_tasks
from .models import ReportSourceBatch, ReportTask, VehicleMonitorDailyRow
from .services import ReportingError


class FiveSourceReportTaskTests(TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.override = override_settings(REPORT_EXPORT_DIR=self.temp_dir.name, REPORT_EXPORT_RETENTION_DAYS=7)
        self.override.enable()
        self.user = get_user_model().objects.create_user(username="five-source-reporter")
        AssistantProfile.objects.create(user=self.user, display_name="五来源报表管理员", employee_code="FIVE-RPT")
        assign_role(user=self.user, role=RoleAssignment.Role.SYSTEM_ADMIN, assigned_by=self.user)
        self.collector = get_user_model().objects.create_user(username="five-source-collector")
        AssistantProfile.objects.create(user=self.collector, display_name="五来源采集员", employee_code="FIVE-COL")
        assign_role(user=self.collector, role=RoleAssignment.Role.MONITOR_OPERATOR, assigned_by=self.user)
        self.enterprise = EnterpriseScope.objects.create(
            code="TEST-ENTERPRISE", name="测试运输企业", scope_type=EnterpriseScope.ScopeType.ENTERPRISE,
        )
        EnterpriseGrant.objects.create(user=self.user, enterprise=self.enterprise)
        EnterpriseGrant.objects.create(user=self.collector, enterprise=self.enterprise)
        claim_shift(user=self.collector, platform_account_ref="five-source-platform", workstation_id="FIVE-SOURCE-WS")

    def tearDown(self):
        self.override.disable()
        self.temp_dir.cleanup()

    def task(self, report_type=ReportTask.ReportType.VEHICLE_MONITOR_DAILY, start="2026-07-21", end="2026-07-21"):
        return report_tasks.create_report_task(
            actor=self.user, report_type=report_type, period_start=start, period_end=end,
        )

    def arm(self, task):
        token = "test-report-lease-token"
        task.claimed_by = self.collector
        task.platform_account_ref = "five-source-platform"
        task.device_id = "FIVE-SOURCE-WS"
        task.lease_token_hash = hashlib.sha256(token.encode()).hexdigest()
        task.lease_expires_at = timezone.now() + timedelta(minutes=10)
        task.status = ReportTask.Status.FETCHING
        task.save(update_fields=["claimed_by", "platform_account_ref", "device_id", "lease_token_hash", "lease_expires_at", "status", "updated_at"])
        return token

    def upload(self, task, token, source, rows, *, page=1):
        batch = task.source_batches.get(source_type=source)
        signature = report_tasks.rows_field_signature(rows)
        report_tasks.upload_source_page(
            actor=self.collector, task=task, source_type=source, page_number=page,
            query_hash=batch.query_hash, field_signature=signature, rows=rows,
            device_id="FIVE-SOURCE-WS", lease_token=token,
        )
        report_tasks.complete_source_batch(
            actor=self.collector, task=task, source_type=source, total_pages=1,
            total_rows=len(rows), field_signature=signature,
            device_id="FIVE-SOURCE-WS", lease_token=token,
        )

    def base(self, vehicle_id, plate, last_location):
        return {
            "vehicleId": vehicle_id, "plate": plate, "enterpriseId": self.enterprise.code,
            "enterpriseName": self.enterprise.name, "vehicleStatus": "OPERATING",
            "lastLocationTime": last_location,
        }

    def track(self, vehicle_id, plate, mileage, completeness):
        return {
            "vehicleId": vehicle_id, "plate": plate, "enterpriseId": self.enterprise.code,
            "enterpriseName": self.enterprise.name, "totalMileage": mileage, "completeness": completeness,
        }

    def test_task_freezes_required_sources_and_blocks_unverified_contracts(self):
        task = self.task()
        self.assertEqual(task.required_source_types, ["VEHICLE_BASE_INFO", "TRACK_COMPLETENESS"])
        self.assertEqual(task.source_batches.count(), 2)
        with self.assertRaises(ReportingError) as caught:
            report_tasks.claim_report_task(actor=self.collector, task=task, device_id="FIVE-SOURCE-WS")
        self.assertEqual(caught.exception.code, "REPORT_CONTRACT_UNVERIFIED")

    def test_page_upload_is_idempotent_and_rejects_credentials_or_conflicts(self):
        task = self.task()
        token = self.arm(task)
        rows = [self.base("vehicle-001", "模拟车A01", "2026-07-21 08:00:00")]
        batch = task.source_batches.get(source_type="VEHICLE_BASE_INFO")
        signature = report_tasks.rows_field_signature(rows)
        first, created = report_tasks.upload_source_page(
            actor=self.collector, task=task, source_type="VEHICLE_BASE_INFO", page_number=1,
            query_hash=batch.query_hash, field_signature=signature, rows=rows,
            device_id="FIVE-SOURCE-WS", lease_token=token,
        )
        same, created_again = report_tasks.upload_source_page(
            actor=self.collector, task=task, source_type="VEHICLE_BASE_INFO", page_number=1,
            query_hash=batch.query_hash, field_signature=signature, rows=rows,
            device_id="FIVE-SOURCE-WS", lease_token=token,
        )
        self.assertEqual(first.pk, same.pk)
        self.assertTrue(created)
        self.assertFalse(created_again)
        conflict = [self.base("vehicle-002", "模拟车A02", "2026-07-21 09:00:00")]
        with self.assertRaises(ReportingError) as caught:
            report_tasks.upload_source_page(
                actor=self.collector, task=task, source_type="VEHICLE_BASE_INFO", page_number=1,
                query_hash=batch.query_hash, field_signature=report_tasks.rows_field_signature(conflict), rows=conflict,
                device_id="FIVE-SOURCE-WS", lease_token=token,
            )
        self.assertEqual(caught.exception.code, "REPORT_PAGE_IDEMPOTENCY_CONFLICT")
        with self.assertRaises(ReportingError) as sensitive:
            report_tasks.upload_source_page(
                actor=self.collector, task=task, source_type="VEHICLE_BASE_INFO", page_number=2,
                query_hash=batch.query_hash, field_signature="x" * 64,
                rows=[{**rows[0], "authorization": "Bearer forbidden"}],
                device_id="FIVE-SOURCE-WS", lease_token=token,
            )
        self.assertEqual(sensitive.exception.code, "SENSITIVE_FIELD_REJECTED")
        self.assertEqual(
            report_tasks.rows_field_signature(rows),
            report_tasks.rows_field_signature([rows[0], {**rows[0], "lastLocationTime": None}]),
        )

    def test_batch_completion_detects_missing_pages(self):
        task = self.task()
        token = self.arm(task)
        rows = [self.base("vehicle-001", "模拟车A01", "2026-07-21 08:00:00")]
        batch = task.source_batches.get(source_type="VEHICLE_BASE_INFO")
        signature = report_tasks.rows_field_signature(rows)
        report_tasks.upload_source_page(
            actor=self.collector, task=task, source_type="VEHICLE_BASE_INFO", page_number=2,
            query_hash=batch.query_hash, field_signature=signature, rows=rows,
            device_id="FIVE-SOURCE-WS", lease_token=token,
        )
        batch, problems = report_tasks.complete_source_batch(
            actor=self.collector, task=task, source_type="VEHICLE_BASE_INFO", total_pages=2, total_rows=1,
            field_signature=signature, device_id="FIVE-SOURCE-WS", lease_token=token,
        )
        self.assertEqual(batch.status, ReportSourceBatch.Status.INVALID)
        self.assertIn("分页不连续", problems)
        task.refresh_from_db()
        self.assertEqual(task.status, ReportTask.Status.DATA_INCOMPLETE)

    def test_vehicle_daily_rules_create_formal_filtered_and_anomaly_rows(self):
        task = self.task()
        token = self.arm(task)
        base_rows = [
            self.base("vehicle-online-zero", "模拟车A01", "2026-07-21 08:00:00"),
            self.base("vehicle-offline", "模拟车A02", "2026-07-20 08:00:00"),
            self.base("vehicle-empty", "模拟车A03", ""),
            self.base("vehicle-future", "模拟车A04", "2026-07-22 08:00:00"),
        ]
        track_rows = [
            self.track("vehicle-online-zero", "模拟车A01", "0", "100%"),
            self.track("vehicle-unmatched", "模拟车A05", "12.5", "90%"),
        ]
        self.upload(task, token, "VEHICLE_BASE_INFO", base_rows)
        self.upload(task, token, "TRACK_COMPLETENESS", track_rows)
        task = report_tasks.finalize_report_task(actor=self.collector, task=task)
        counts = Counter(VehicleMonitorDailyRow.objects.filter(task=task).values_list("row_kind", flat=True))
        self.assertEqual(counts[VehicleMonitorDailyRow.RowKind.FORMAL], 2)
        self.assertEqual(counts[VehicleMonitorDailyRow.RowKind.FILTERED], 1)
        self.assertEqual(counts[VehicleMonitorDailyRow.RowKind.ANOMALY], 2)
        zero = next(row for row in VehicleMonitorDailyRow.objects.filter(task=task) if row.data.get("vehicleId") == "vehicle-online-zero")
        self.assertEqual(zero.trajectory_status, "在线无轨迹/未行驶")
        self.assertGreater(task.critical_issue_count, 0)
        with self.assertRaises(ReportingError) as caught:
            report_tasks.review_report_task(actor=self.user, task=task, approve=True)
        self.assertEqual(caught.exception.code, "REPORT_CRITICAL_ISSUES_UNRESOLVED")

    def test_clean_vehicle_task_exports_four_sheet_workbook_and_bundle(self):
        task = self.task()
        token = self.arm(task)
        self.upload(task, token, "VEHICLE_BASE_INFO", [
            self.base("vehicle-online", "模拟车A01", "2026-07-21 08:00:00"),
            self.base("vehicle-offline", "模拟车A02", "2026-07-20 08:00:00"),
        ])
        self.upload(task, token, "TRACK_COMPLETENESS", [self.track("vehicle-online", "模拟车A01", "20", "100%")])
        task = report_tasks.finalize_report_task(actor=self.collector, task=task)
        self.assertEqual(task.critical_issue_count, 0)
        task = report_tasks.review_report_task(actor=self.user, task=task, approve=True, note="黄金样本通过")
        snapshot = task.snapshots.get()
        content = report_tasks.snapshot_xlsx(snapshot)
        workbook = load_workbook(BytesIO(content), data_only=False)
        self.assertEqual(workbook.sheetnames, [snapshot.enterprise.name[:20] + "车辆动态监控日报表", "明细审计", "过滤清单", "异常清单"])
        main = workbook.worksheets[0]
        self.assertEqual(main["B6"].value, "模拟车A01")
        for column in ["C", "D", "E", "I", "J", "K", "L"]:
            self.assertIsNone(main[f"{column}6"].value)
        workbook.close()
        job = report_tasks.create_task_bundle_export(actor=self.user, task=task, purpose="测试批量下载")
        with ZipFile(job.file_path) as archive:
            self.assertIn("manifest.json", archive.namelist())
            self.assertIn("_全局异常清单.xlsx", archive.namelist())

    def test_duplicate_and_missing_vehicle_ids_are_not_silently_merged(self):
        task = self.task()
        token = self.arm(task)
        self.upload(task, token, "VEHICLE_BASE_INFO", [
            self.base("duplicate-id", "模拟车D01", "2026-07-21 08:00:00"),
            self.base("duplicate-id", "模拟车D02", "2026-07-21 09:00:00"),
            self.base("", "模拟车M01", "2026-07-21 10:00:00"),
            self.base("", "模拟车M02", "2026-07-21 11:00:00"),
        ])
        self.upload(task, token, "TRACK_COMPLETENESS", [])
        task = report_tasks.finalize_report_task(actor=self.collector, task=task)
        rows = list(VehicleMonitorDailyRow.objects.filter(task=task))
        self.assertEqual(len(rows), 4)
        self.assertTrue(all(row.row_kind == VehicleMonitorDailyRow.RowKind.ANOMALY for row in rows))
        self.assertEqual(len({row.vehicle_key_hash for row in rows}), 4)

    def test_alarm_sources_remain_separate_and_use_fixed_three_sheet_template(self):
        task = self.task(ReportTask.ReportType.ALARM_DAILY)
        token = self.arm(task)
        for source in report_tasks.ALARM_SOURCES:
            values = {header: f"测试-{source}-{index}" for index, header in enumerate(report_tasks.ALARM_HEADERS[source], 1)}
            values["序号"] = 1
            self.upload(task, token, source, [{
                "enterpriseId": self.enterprise.code, "enterpriseName": self.enterprise.name, "values": values,
            }])
        task = report_tasks.finalize_report_task(actor=self.collector, task=task)
        self.assertEqual(task.validation_summary["sourceCounts"]["ALARM_CENTER"]["valid"], 1)
        task = report_tasks.review_report_task(actor=self.user, task=task, approve=True)
        workbook = load_workbook(BytesIO(report_tasks.snapshot_xlsx(task.snapshots.get())), data_only=False)
        self.assertEqual(workbook.sheetnames, ["处置率报表", "处理率报表", "报警查询报表"])
        self.assertEqual(workbook["处置率报表"]["B2"].value, "机构")
        self.assertEqual(workbook["处理率报表"]["K2"].value, "处理率")
        self.assertEqual(workbook["报警查询报表"]["C2"].value, "报警ID")
        workbook.close()
