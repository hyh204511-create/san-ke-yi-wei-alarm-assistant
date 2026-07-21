import calendar
import hashlib
import json
import re
import secrets
import uuid
from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta
from io import BytesIO
from pathlib import Path
from urllib.parse import urlparse

from django.conf import settings
from django.db import IntegrityError, transaction
from django.db.models import Max
from django.utils import timezone
from django.utils.dateparse import parse_datetime

from apps.disposals.models import DisposalCase
from apps.governance.models import AuditEvent, DeviceRegistration, EnterpriseScope, VoiceInteractionPolicy
from apps.governance.services import (
    GovernanceError, active_roles, active_shift_for_user, enterprise_scope_for_user, enterprise_scope_ids_for_user,
    permissions_for_roles, require_permission, resolve_enterprise_for_user, voice_interaction_policy,
)

from .models import ActionLease, AlarmFact, CaptureSource, DutyNotification, ExportJob, ReportSnapshot, VoiceInteractionEvidence


class ReportingError(Exception):
    def __init__(self, message, code="REPORTING_ERROR", status=400):
        super().__init__(message)
        self.code = code
        self.status = status


ACTION_RESULT_CODES = {"EXECUTING", "SUCCEEDED", "FAILED", "UNKNOWN", "BLOCKED", "MANUAL_REQUIRED"}
ACTION_FAILURE_CODES = {"FAILED", "UNKNOWN", "BLOCKED", "MANUAL_REQUIRED"}
SAFE_ACTION_RESULT_KEYS = {
    "receiptRef", "errorCode", "latencyMs", "attemptNumber", "simulated", "terminalTts",
    "playbackStarted", "platformHttpStatus", "messageCode",
}
VOICE_EVIDENCE_KEYS = {
    "leaseToken", "deviceId", "audioSha256", "durationMs", "recordedStartedAt", "recordedEndedAt",
    "audioRef", "transcript", "transcriptEngine", "transcriptConfidence", "source",
    "consentAccepted", "consentReference",
}
VOICE_TRANSCRIPT_KEYS = {"transcriptText", "language", "utteranceDetected", "confidence", "engine", "source"}


def require_reporting_permission(actor, permission, *, require_shift=False):
    try:
        require_permission(actor, permission)
    except GovernanceError as exc:
        raise ReportingError(str(exc), exc.code, exc.status) from exc
    if require_shift and not active_shift_for_user(actor):
        raise ReportingError("请先认领当前值班班次", "ACTIVE_SHIFT_REQUIRED", 409)


def require_voice_evidence_permission(actor, *, include_transcript=False):
    """Enforce the distinction between evidence metadata and transcript review.

    Unit users may see that evidence exists, while only evidence reviewers and
    system administrators may read the encrypted transcript itself.
    """
    permissions = set(permissions_for_roles(active_roles(actor)))
    required = "evidence.review" if include_transcript else "evidence.view"
    if required not in permissions:
        label = "转写内容" if include_transcript else "摘要"
        raise ReportingError(f"当前角色不能查看语音证据{label}", "PERMISSION_DENIED", 403)


def parse_platform_datetime(value):
    if not value:
        return None
    parsed = parse_datetime(str(value).replace(" ", "T"))
    if not parsed:
        try:
            parsed = datetime.strptime(str(value), "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return None
    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return parsed


def require_scope(actor, enterprise):
    if enterprise.pk not in enterprise_scope_ids_for_user(actor):
        raise ReportingError("无权访问该企业报表", "ENTERPRISE_SCOPE_DENIED", 403)


def audit(actor, event_type, object_type, object_id, detail):
    AuditEvent.objects.create(
        actor=actor, event_type=event_type, object_type=object_type, object_id=str(object_id),
        role_snapshot=list(actor.assistant_roles.filter(is_active=True).values_list("role", flat=True)),
        enterprise_scope_snapshot=enterprise_scope_for_user(actor), detail=detail,
    )


def merge_ingestion_provenance(existing, actor, seen_at):
    shift = active_shift_for_user(actor)
    source = {
        "userId": str(actor.pk),
        "shiftId": str(shift.public_id),
        "platformAccountRef": shift.platform_account_ref,
        "workstationId": shift.workstation_id,
        "firstSeenAt": seen_at.isoformat(),
        "lastSeenAt": seen_at.isoformat(),
    }
    rows = [dict(item) for item in (existing or []) if isinstance(item, dict)]
    key = (source["userId"], source["shiftId"])
    for row in rows:
        if (str(row.get("userId")), str(row.get("shiftId"))) == key:
            row["lastSeenAt"] = source["lastSeenAt"]
            return rows[-100:]
    rows.append(source)
    return rows[-100:]


def alarm_business_fingerprint(event):
    alarm_id = str(event.get("alarmId") or "").strip()
    if alarm_id:
        identity = ["ALARM_ID", alarm_id]
    else:
        identity = [
            "BUSINESS", str(event.get("vehicleId") or event.get("vehicleNo") or ""),
            str(event.get("alarmTypeKey") or event.get("alarmName") or ""),
            str(event.get("alarmTime") or event.get("discoveredAt") or ""),
            str(event.get("companyId") or event.get("companyName") or ""),
        ]
    return hashlib.sha256(json.dumps(identity, ensure_ascii=False, separators=(",", ":")).encode("utf-8")).hexdigest()


def normalized_capture_source(actor, event, source, seen_at):
    shift = active_shift_for_user(actor)
    source = source if isinstance(source, dict) else {}
    allowed = {"captureId", "deviceId", "platformAccountRef", "extensionVersion", "endpoint", "capturedAt"}
    if set(source) - allowed:
        raise ReportingError("采集来源包含未允许字段", "INVALID_CAPTURE_SOURCE", 422)
    platform_ref = str(source.get("platformAccountRef") or shift.platform_account_ref)
    if platform_ref != shift.platform_account_ref:
        raise ReportingError("采集来源的平台账号引用与当前班次不一致", "PLATFORM_ACCOUNT_MISMATCH", 409)
    device_id = str(source.get("deviceId") or shift.workstation_id).strip()
    if not device_id or len(device_id) > 120:
        raise ReportingError("采集设备标识无效", "INVALID_DEVICE_ID", 422)
    captured_at = parse_platform_datetime(source.get("capturedAt")) or seen_at
    capture_id = str(source.get("captureId") or event.get("captureId") or "").strip()
    if not capture_id:
        capture_id = hashlib.sha256(f"{device_id}|{event.get('eventId')}|{captured_at.isoformat()}".encode()).hexdigest()
    if len(capture_id) > 160:
        raise ReportingError("采集幂等标识无效", "INVALID_CAPTURE_ID", 422)
    endpoint_value = str(source.get("endpoint") or event.get("rawEndpoint") or "")
    parsed_endpoint = urlparse(endpoint_value)
    endpoint = (parsed_endpoint.path if parsed_endpoint.scheme or parsed_endpoint.netloc else endpoint_value.split("?", 1)[0])[:300]
    payload_hash = hashlib.sha256(json.dumps(event, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode()).hexdigest()
    return {
        "capture_id": capture_id, "device_id": device_id, "platform_account_ref": platform_ref,
        "extension_version": str(source.get("extensionVersion") or "")[:40], "endpoint": endpoint,
        "captured_at": captured_at, "payload_hash": payload_hash,
    }


def capture_defaults(capture, actor):
    return {key: value for key, value in capture.items() if key not in {"device_id", "capture_id"}} | {"ingested_by": actor}


@transaction.atomic
def upsert_alarm_fact(*, actor, event, decision, action, source=None):
    require_reporting_permission(actor, "alarm.view", require_shift=True)
    if not isinstance(event, dict) or not isinstance(decision, dict):
        raise ReportingError("报警和判断必须是JSON对象", "INVALID_PAYLOAD", 422)
    event_id = str(event.get("eventId") or "").strip()
    if not event_id or len(event_id) > 160:
        raise ReportingError("报警事件标识无效", "INVALID_EVENT_ID", 422)
    try:
        enterprise = resolve_enterprise_for_user(actor, event.get("companyId"), event.get("companyName"))
    except GovernanceError as exc:
        raise ReportingError(str(exc), exc.code, exc.status) from exc
    seen_at = parse_platform_datetime(event.get("updatedAt") or event.get("discoveredAt")) or timezone.now()
    alarm_time = parse_platform_datetime(event.get("alarmTime"))
    fingerprint = alarm_business_fingerprint(event)
    capture = normalized_capture_source(actor, event, source, seen_at)
    completion = event.get("completionAssessment") if isinstance(event.get("completionAssessment"), dict) else {}
    fact = AlarmFact.objects.select_for_update().filter(event_id=event_id).first()
    if not fact:
        fact = AlarmFact.objects.select_for_update().filter(business_fingerprint=fingerprint).first()
    if fact and fact.enterprise_id != enterprise.pk:
        raise ReportingError("同一报警的企业归属发生冲突", "ENTERPRISE_CONFLICT", 409)
    values = {
        "alarm_id": str(event.get("alarmId") or "")[:160], "business_fingerprint": fingerprint, "enterprise": enterprise,
        "company_name_snapshot": str(event.get("companyName") or enterprise.name)[:200], "source_kind": str(event.get("sourceKind") or "OTHER")[:30],
        "alarm_name": str(event.get("alarmName") or "")[:200], "alarm_time": alarm_time,
        "vehicle_id": str(event.get("vehicleId") or "")[:160], "vehicle_no": str(event.get("vehicleNo") or "")[:100],
        "final_state": str(event.get("state") or "")[:40], "event_snapshot": event, "decision_snapshot": decision,
        "completion_status": str(completion.get("status") or "UNKNOWN_MANUAL")[:30],
        "completion_source": str(completion.get("source") or "MANUAL_CONFIRMATION")[:30],
        "completion_manual_required": bool(completion.get("manualRequired", True)),
        "completion_reason": str(completion.get("reason") or "")[:500],
        "action_snapshot": action if isinstance(action, dict) else {}, "last_seen_at": seen_at,
    }
    if fact:
        values["ingestion_provenance"] = merge_ingestion_provenance(fact.ingestion_provenance, actor, seen_at)
        changed = any(getattr(fact, key) != value for key, value in values.items())
        if not changed:
            CaptureSource.objects.get_or_create(fact=fact, device_id=capture["device_id"], capture_id=capture["capture_id"], defaults=capture_defaults(capture, actor))
            notify_from_action_snapshot(actor=actor, fact=fact, action=action)
            return fact, False
        for key, value in values.items(): setattr(fact, key, value)
        fact.save()
        CaptureSource.objects.get_or_create(fact=fact, device_id=capture["device_id"], capture_id=capture["capture_id"], defaults=capture_defaults(capture, actor))
        notify_from_action_snapshot(actor=actor, fact=fact, action=action)
        return fact, False
    created = True
    try:
        with transaction.atomic():
            fact = AlarmFact.objects.create(
                event_id=event_id, first_seen_at=seen_at, ingested_by=actor,
                ingestion_provenance=merge_ingestion_provenance([], actor, seen_at), **values,
            )
    except IntegrityError:
        created = False
        fact = AlarmFact.objects.select_for_update().filter(business_fingerprint=fingerprint).first() or AlarmFact.objects.select_for_update().get(event_id=event_id)
        if fact.enterprise_id != enterprise.pk:
            raise ReportingError("同一报警的企业归属发生冲突", "ENTERPRISE_CONFLICT", 409)
    CaptureSource.objects.get_or_create(fact=fact, device_id=capture["device_id"], capture_id=capture["capture_id"], defaults=capture_defaults(capture, actor))
    notify_from_action_snapshot(actor=actor, fact=fact, action=action)
    return fact, created


def _lease_token_hash(token):
    return hashlib.sha256(str(token).encode("utf-8")).hexdigest()


@transaction.atomic
def acquire_action_lease(*, actor, fact, device_id, action_type, duration_seconds=120, mode="LIVE", require_registered_device=False):
    require_reporting_permission(actor, "action.execute", require_shift=True)
    fact = AlarmFact.objects.select_for_update().get(pk=fact.pk)
    device_id = str(device_id or "").strip()
    if not device_id or len(device_id) > 120:
        raise ReportingError("动作设备标识无效", "INVALID_DEVICE_ID", 422)
    device = DeviceRegistration.objects.filter(device_id=device_id, user=actor, is_active=True).first()
    if require_registered_device and not device:
        raise ReportingError("动作设备尚未登记", "DEVICE_NOT_REGISTERED", 409)
    if device and str(mode or "LIVE").upper() == "LIVE" and device.platform_identity_status != "VERIFIED":
        raise ReportingError("省平台账号身份尚未核验，禁止真实动作", "PLATFORM_IDENTITY_REQUIRED", 409)
    action_type = str(action_type or "").strip().upper()
    action_type = {"TEXT": "TEXT_TTS", "VOICE": "VOICE_INTERCOM"}.get(action_type, action_type)
    if action_type not in {"TEXT_TTS", "VOICE_INTERCOM"}:
        raise ReportingError("动作类型必须是TEXT_TTS或VOICE_INTERCOM", "INVALID_ACTION_TYPE", 422)
    now = timezone.now()
    active_statuses = [ActionLease.Status.ACTIVE, ActionLease.Status.EXECUTING]
    ActionLease.objects.filter(fact=fact, action_type=action_type, status__in=active_statuses, expires_at__lte=now).update(status=ActionLease.Status.EXPIRED, finished_at=now)
    if ActionLease.objects.filter(fact=fact, action_type=action_type, status__in=active_statuses).exists():
        raise ReportingError("该报警已有设备取得动作租约", "ACTION_LEASE_CONFLICT", 409)
    raw_token = secrets.token_urlsafe(32)
    lease = ActionLease.objects.create(
        fact=fact, actor=actor, device_id=str(device_id)[:120], action_type=str(action_type)[:60],
        lease_token_hash=_lease_token_hash(raw_token),
        expires_at=now + timedelta(seconds=max(30, min(int(duration_seconds), 600))),
    )
    # The raw token is returned once to the caller and is never persisted or logged.
    lease._plain_token = raw_token
    return lease


def action_lease_payload(lease):
    return {
        "leaseId": str(lease.public_id),
        "actionType": lease.action_type,
        "status": lease.status,
        "expiresAt": lease.expires_at.isoformat(),
        "leaseToken": getattr(lease, "_plain_token", None),
    }


def fact_for_event(*, actor, event_id):
    require_reporting_permission(actor, "alarm.view", require_shift=True)
    event_id = str(event_id or "").strip()
    if not event_id or len(event_id) > 160:
        raise ReportingError("报警事件标识无效", "INVALID_EVENT_ID", 422)
    fact = AlarmFact.objects.select_related("enterprise").filter(event_id=event_id).first()
    if not fact:
        raise ReportingError("报警事实不存在，请先完成事件入库", "ALARM_FACT_NOT_FOUND", 404)
    require_scope(actor, fact.enterprise)
    return fact


def _safe_action_result(payload):
    result = payload.get("result") if isinstance(payload.get("result"), dict) else {}
    if set(result) - SAFE_ACTION_RESULT_KEYS:
        raise ReportingError("动作回执包含未允许字段", "INVALID_ACTION_RESULT", 422)
    safe = {}
    for key, value in result.items():
        if key == "receiptRef":
            safe[key] = str(value)[:160]
        elif key == "errorCode" or key == "messageCode":
            safe[key] = str(value)[:80]
        elif key == "latencyMs" or key == "attemptNumber" or key == "platformHttpStatus":
            try:
                safe[key] = max(0, min(int(value), 300000))
            except (TypeError, ValueError) as exc:
                raise ReportingError("动作回执数字字段无效", "INVALID_ACTION_RESULT", 422) from exc
        elif key in {"simulated", "terminalTts", "playbackStarted"}:
            if not isinstance(value, bool):
                raise ReportingError("动作回执布尔字段无效", "INVALID_ACTION_RESULT", 422)
            safe[key] = value
    return safe


def _notification_message(result_code, action_type):
    labels = {
        "FAILED": "明确失败",
        "UNKNOWN": "结果未知",
        "BLOCKED": "被安全闸门阻断",
        "MANUAL_REQUIRED": "需要人工接管",
    }
    return f"{action_type} {labels.get(result_code, result_code)}，请当前值班人员查看并处理。"


def ensure_action_notification(*, actor, fact, result_code, action_type, action_id="", lease=None, detail=None):
    result_code = str(result_code or "").upper()
    if result_code not in ACTION_FAILURE_CODES:
        return None
    lease_ref = str(lease.public_id) if lease else ""
    action_ref = str(action_id or lease_ref)
    dedupe_key = hashlib.sha256(f"{fact.event_id}|{action_ref}|{action_type}|{result_code}".encode("utf-8")).hexdigest()
    notification, _ = DutyNotification.objects.get_or_create(
        dedupe_key=dedupe_key,
        defaults={
            "recipient": actor,
            "enterprise": fact.enterprise,
            "action_lease": lease,
            "event_id": fact.event_id,
            "kind": "ACTION_MANUAL_REQUIRED",
            "result_code": result_code,
            "title": "自动动作需要人工处理",
            "message": _notification_message(result_code, action_type),
            "detail": detail or {},
        },
    )
    return notification


def ensure_voice_review_notification(*, actor, fact, action_id="", lease=None, detail=None):
    """Create one post-success review task for an automated voice action."""
    lease_ref = str(lease.public_id) if lease else ""
    action_ref = str(action_id or lease_ref)
    dedupe_key = hashlib.sha256(
        f"{fact.event_id}|{action_ref}|VOICE_INTERCOM|VOICE_REVIEW_REQUIRED".encode("utf-8")
    ).hexdigest()
    notification, _ = DutyNotification.objects.get_or_create(
        dedupe_key=dedupe_key,
        defaults={
            "recipient": actor,
            "enterprise": fact.enterprise,
            "action_lease": lease,
            "event_id": fact.event_id,
            "kind": "VOICE_REVIEW_REQUIRED",
            "result_code": "SUCCEEDED",
            "title": "自动语音已播报，请人工审核",
            "message": "VOICE_INTERCOM 已收到成功回执，请当前值班人员查看动作记录和可用语音证据。",
            "detail": detail or {},
        },
    )
    return notification


def notify_from_action_snapshot(*, actor, fact, action):
    if not isinstance(action, dict):
        return None
    status = str(action.get("status") or "").upper()
    action_type = str(action.get("channelType") or action.get("type") or "ACTION").upper()
    if action_type == "RESPONSE_PLAN":
        action_type = "VOICE_INTERCOM" if any(
            item.get("channelType") == "VOICE" for item in action.get("attempts", []) if isinstance(item, dict)
        ) else "TEXT_TTS"
        statuses = [str(item.get("status") or "").upper() for item in action.get("attempts", []) if isinstance(item, dict)]
        status = next((item for item in ["UNKNOWN", "BLOCKED", "FAILED", "MANUAL_REQUIRED"] if item in statuses), status)
    if status not in ACTION_FAILURE_CODES:
        return None
    return ensure_action_notification(
        actor=actor, fact=fact, result_code=status, action_type=action_type,
        action_id=str(action.get("actionId") or ""),
        detail={"blockers": [str(item)[:200] for item in action.get("blockers", [])[:20]]},
    )


@transaction.atomic
def record_action_result(*, actor, payload):
    require_reporting_permission(actor, "action.execute", require_shift=True)
    if not isinstance(payload, dict):
        raise ReportingError("动作回执必须是JSON对象", "INVALID_ACTION_RESULT", 422)
    allowed = {"leaseId", "leaseToken", "deviceId", "resultCode", "result", "actionId"}
    if set(payload) - allowed:
        raise ReportingError("动作回执包含未允许字段", "INVALID_ACTION_RESULT", 422)
    try:
        lease_id = uuid.UUID(str(payload.get("leaseId")))
    except (TypeError, ValueError, AttributeError) as exc:
        raise ReportingError("动作租约标识无效", "INVALID_ACTION_LEASE", 422) from exc
    lease = ActionLease.objects.select_for_update().select_related("fact__enterprise").filter(public_id=lease_id).first()
    if not lease:
        raise ReportingError("动作租约不存在", "ACTION_LEASE_NOT_FOUND", 404)
    if lease.actor_id != actor.pk or lease.device_id != str(payload.get("deviceId") or ""):
        raise ReportingError("动作租约不属于当前用户或设备", "ACTION_LEASE_OWNER_CONFLICT", 403)
    if not secrets.compare_digest(lease.lease_token_hash, _lease_token_hash(payload.get("leaseToken") or "")):
        raise ReportingError("动作租约令牌无效", "ACTION_LEASE_TOKEN_INVALID", 403)
    result_code = str(payload.get("resultCode") or "").upper()
    if result_code not in ACTION_RESULT_CODES:
        raise ReportingError("动作结果码无效", "INVALID_ACTION_RESULT", 422)
    if lease.status not in {ActionLease.Status.ACTIVE, ActionLease.Status.EXECUTING}:
        if lease.result_code == result_code:
            return lease, False
        raise ReportingError("动作租约已经结束", "ACTION_LEASE_CLOSED", 409)
    now = timezone.now()
    result_payload = _safe_action_result(payload)
    lease.result_code = result_code
    lease.result_payload = result_payload
    lease.last_attempt_at = now
    if result_code == "EXECUTING":
        lease.status = ActionLease.Status.EXECUTING
        lease.started_at = lease.started_at or now
    else:
        lease.status = {
            "SUCCEEDED": ActionLease.Status.COMPLETED,
            "FAILED": ActionLease.Status.FAILED,
            "UNKNOWN": ActionLease.Status.UNKNOWN,
            "BLOCKED": ActionLease.Status.BLOCKED,
            "MANUAL_REQUIRED": ActionLease.Status.MANUAL_REQUIRED,
        }[result_code]
        lease.finished_at = now
    lease.save(update_fields=["result_code", "result_payload", "last_attempt_at", "started_at", "status", "finished_at", "updated_at"])
    fact = AlarmFact.objects.select_for_update().get(pk=lease.fact_id)
    action_snapshot = dict(fact.action_snapshot or {}) if isinstance(fact.action_snapshot, dict) else {}
    action_snapshot["serverResult"] = {
        "leaseId": str(lease.public_id), "actionType": lease.action_type,
        "resultCode": result_code, "result": result_payload, "recordedAt": now.isoformat(),
    }
    fact.action_snapshot = action_snapshot
    fact.save(update_fields=["action_snapshot", "updated_at"])
    audit(actor, "ACTION_RESULT_RECORDED", "ACTION_LEASE", lease.public_id, {"resultCode": result_code, "actionType": lease.action_type})
    notification = ensure_action_notification(
        actor=actor, fact=lease.fact, result_code=result_code, action_type=lease.action_type,
        action_id=payload.get("actionId") or "", lease=lease, detail={"result": result_payload},
    )
    if result_code == "SUCCEEDED" and lease.action_type == "VOICE_INTERCOM":
        notification = ensure_voice_review_notification(
            actor=actor, fact=lease.fact, action_id=payload.get("actionId") or "", lease=lease,
            detail={"result": result_payload},
        )
    return lease, bool(notification)


def list_notifications(*, actor, include_acknowledged=False, limit=100):
    require_reporting_permission(actor, "alarm.view")
    queryset = DutyNotification.objects.filter(recipient=actor).select_related("enterprise")
    if not include_acknowledged:
        queryset = queryset.filter(status=DutyNotification.Status.UNREAD)
    return list(queryset[:max(1, min(int(limit), 200))])


@transaction.atomic
def acknowledge_notification(*, actor, notification):
    require_reporting_permission(actor, "alarm.view")
    notification = DutyNotification.objects.select_for_update().get(pk=notification.pk)
    if notification.recipient_id != actor.pk:
        raise ReportingError("只能确认发给当前值班人员的通知", "NOTIFICATION_OWNER_CONFLICT", 403)
    if notification.status == DutyNotification.Status.UNREAD:
        notification.status = DutyNotification.Status.ACKNOWLEDGED
        notification.acknowledged_by = actor
        notification.acknowledged_at = timezone.now()
        notification.save(update_fields=["status", "acknowledged_by", "acknowledged_at", "updated_at"])
        audit(actor, "DUTY_NOTIFICATION_ACKNOWLEDGED", "DUTY_NOTIFICATION", notification.public_id, {"kind": notification.kind})
    return notification


def _parse_optional_timestamp(value, field_name):
    if not value:
        return None
    parsed = parse_platform_datetime(value)
    if not parsed:
        raise ReportingError(f"{field_name}时间格式无效", "INVALID_VOICE_EVIDENCE", 422)
    return parsed


def _validate_voice_transcript(value, *, allow_empty=False):
    if value is None:
        return None
    if not isinstance(value, dict) or set(value) - {"text", "language", "utteranceDetected", "confidence", "engine"}:
        raise ReportingError("转写结果包含未允许字段", "INVALID_VOICE_TRANSCRIPT", 422)
    text = str(value.get("text") or "")
    if not allow_empty and not text.strip():
        raise ReportingError("转写文本不能为空", "INVALID_VOICE_TRANSCRIPT", 422)
    if len(text) > 2000:
        raise ReportingError("转写文本不能超过2000字符", "INVALID_VOICE_TRANSCRIPT", 422)
    language = str(value.get("language") or "zh-CN")[:20]
    engine = str(value.get("engine") or "")[:80]
    raw_detected = value.get("utteranceDetected")
    detected = bool(text.strip()) if raw_detected is None else raw_detected
    if not isinstance(detected, bool):
        raise ReportingError("是否检测到司机语音必须是布尔值", "INVALID_VOICE_TRANSCRIPT", 422)
    confidence = value.get("confidence")
    if confidence is not None:
        try:
            confidence = float(confidence)
        except (TypeError, ValueError) as exc:
            raise ReportingError("转写置信度无效", "INVALID_VOICE_TRANSCRIPT", 422) from exc
        if confidence < 0 or confidence > 1:
            raise ReportingError("转写置信度必须在0至1之间", "INVALID_VOICE_TRANSCRIPT", 422)
    return {"text": text, "language": language, "utteranceDetected": detected, "confidence": confidence, "engine": engine}


def _validate_voice_consent(payload, policy):
    accepted = payload.get("consentAccepted")
    if accepted is not None and not isinstance(accepted, bool):
        raise ReportingError("录音同意标记必须是布尔值", "INVALID_VOICE_CONSENT", 422)
    reference = str(payload.get("consentReference") or "").strip()
    if len(reference) > 200 or re.search(r"[\r\n]", reference):
        raise ReportingError("录音同意引用无效", "INVALID_VOICE_CONSENT", 422)
    if policy.consent_required and accepted is not True:
        raise ReportingError("录音前必须取得司机同意", "VOICE_CONSENT_REQUIRED", 409)
    if policy.consent_required and not reference:
        raise ReportingError("录音同意必须提供可审计引用", "VOICE_CONSENT_REFERENCE_REQUIRED", 422)
    return {
        "accepted": accepted is True,
        "referenceHash": hashlib.sha256(reference.encode("utf-8")).hexdigest() if reference else "",
    }


def voice_evidence_data(record, *, include_transcript=False):
    data = {
        "evidenceId": str(record.public_id),
        "eventId": record.event_id,
        "enterpriseId": str(record.enterprise.public_id),
        "status": record.status,
        "transcriptionStatus": record.transcription_status,
        "policyVersion": record.policy_version,
        "audioSha256": record.audio_sha256 or None,
        "audioDurationMs": record.audio_duration_ms,
        "recordedStartedAt": record.recorded_started_at.isoformat() if record.recorded_started_at else None,
        "recordedEndedAt": record.recorded_ended_at.isoformat() if record.recorded_ended_at else None,
        "retentionUntil": record.retention_until.isoformat(),
    }
    if include_transcript:
        data["transcript"] = record.transcript or {}
    return data


def _locked_voice_lease(*, actor, lease_id, device_id, lease_token):
    try:
        parsed_id = uuid.UUID(str(lease_id))
    except (TypeError, ValueError, AttributeError) as exc:
        raise ReportingError("语音动作租约标识无效", "INVALID_ACTION_LEASE", 422) from exc
    lease = ActionLease.objects.select_for_update().select_related("fact__enterprise").filter(public_id=parsed_id).first()
    if not lease:
        raise ReportingError("语音动作租约不存在", "ACTION_LEASE_NOT_FOUND", 404)
    if lease.action_type != "VOICE_INTERCOM" or lease.actor_id != actor.pk or lease.device_id != str(device_id or ""):
        raise ReportingError("语音证据与动作租约不匹配", "VOICE_LEASE_OWNER_CONFLICT", 403)
    if not secrets.compare_digest(lease.lease_token_hash or "", _lease_token_hash(lease_token or "")):
        raise ReportingError("语音动作租约令牌无效", "ACTION_LEASE_TOKEN_INVALID", 403)
    if lease.status in {ActionLease.Status.ACTIVE, ActionLease.Status.EXECUTING}:
        raise ReportingError("语音动作尚未结束，不能登记司机回传证据", "VOICE_ACTION_NOT_FINISHED", 409)
    return lease


@transaction.atomic
def register_voice_evidence(*, actor, lease_id, payload):
    require_reporting_permission(actor, "action.execute", require_shift=True)
    if not isinstance(payload, dict) or set(payload) - VOICE_EVIDENCE_KEYS:
        raise ReportingError("语音证据请求包含未允许字段", "INVALID_VOICE_EVIDENCE", 422)
    policy = voice_interaction_policy()
    if not policy.enabled or not policy.record_driver_audio:
        raise ReportingError("当前未授权录制司机回传语音", "VOICE_RECORDING_DISABLED", 409)
    # Raw audio bytes are intentionally not accepted by this JSON endpoint.
    if any(key in payload for key in {"audioBase64", "audioBytes", "content", "blob"}):
        raise ReportingError("原始音频不能通过动作回执接口上传", "VOICE_AUDIO_BYTES_NOT_ACCEPTED", 422)
    lease = _locked_voice_lease(
        actor=actor, lease_id=lease_id, device_id=payload.get("deviceId"), lease_token=payload.get("leaseToken"),
    )
    consent = _validate_voice_consent(payload, policy)
    audio_hash = str(payload.get("audioSha256") or "").lower()
    if audio_hash and not re.fullmatch(r"[0-9a-f]{64}", audio_hash):
        raise ReportingError("音频摘要必须是SHA-256", "INVALID_VOICE_EVIDENCE", 422)
    try:
        duration = int(payload.get("durationMs") or 0)
    except (TypeError, ValueError) as exc:
        raise ReportingError("音频时长无效", "INVALID_VOICE_EVIDENCE", 422) from exc
    if duration < 0 or duration > 120000:
        raise ReportingError("音频时长必须在0至120秒之间", "INVALID_VOICE_EVIDENCE", 422)
    audio_ref = str(payload.get("audioRef") or "")
    if audio_ref and (len(audio_ref) > 300 or ".." in audio_ref or re.search(r"[\r\n]", audio_ref) or audio_ref.startswith(("http://", "https://"))):
        raise ReportingError("音频引用必须是受控存储引用，不能是公网URL", "INVALID_VOICE_EVIDENCE", 422)
    transcript = _validate_voice_transcript(payload.get("transcript"), allow_empty=True)
    if transcript is not None and not policy.transcribe_driver_audio:
        raise ReportingError("当前未授权司机语音转文字", "VOICE_TRANSCRIPTION_DISABLED", 409)
    if transcript is not None and str(payload.get("source") or "") != "LOCAL_TRANSCRIBER":
        raise ReportingError("登记时的转写结果必须来自受控本地转写适配器", "INVALID_TRANSCRIPTION_SOURCE", 422)
    started_at = _parse_optional_timestamp(payload.get("recordedStartedAt"), "recordedStartedAt")
    ended_at = _parse_optional_timestamp(payload.get("recordedEndedAt"), "recordedEndedAt")
    if started_at and ended_at and ended_at < started_at:
        raise ReportingError("录音结束时间不能早于开始时间", "INVALID_VOICE_EVIDENCE", 422)
    dedupe_key = hashlib.sha256(f"{lease.public_id}|{audio_hash}|{started_at.isoformat() if started_at else ''}".encode("utf-8")).hexdigest()
    record, created = VoiceInteractionEvidence.objects.get_or_create(
        dedupe_key=dedupe_key,
        defaults={
            "fact": lease.fact, "action_lease": lease, "enterprise": lease.fact.enterprise,
            "requested_by": actor, "event_id": lease.fact.event_id, "policy_version": policy.version,
            "status": VoiceInteractionEvidence.Status.TRANSCRIBED if transcript is not None else VoiceInteractionEvidence.Status.CAPTURED,
            "transcription_status": VoiceInteractionEvidence.TranscriptionStatus.READY if transcript is not None else (
                VoiceInteractionEvidence.TranscriptionStatus.PENDING if policy.transcribe_driver_audio else VoiceInteractionEvidence.TranscriptionStatus.NOT_REQUESTED
            ),
            "audio_sha256": audio_hash, "audio_duration_ms": duration,
            "audio_metadata": {
                "audioRef": audio_ref, "source": str(payload.get("source") or "")[:80],
                "consentAccepted": consent["accepted"], "consentReferenceHash": consent["referenceHash"],
            },
            "transcript": transcript or {}, "transcript_engine": (transcript or {}).get("engine", ""),
            "transcript_confidence": (transcript or {}).get("confidence"),
            "recorded_started_at": started_at, "recorded_ended_at": ended_at,
            "retention_until": timezone.now() + timedelta(days=policy.retention_days),
        },
    )
    audit(actor, "VOICE_EVIDENCE_REGISTERED", "VOICE_INTERACTION_EVIDENCE", record.public_id, {
        "eventId": record.event_id, "status": record.status, "transcriptionStatus": record.transcription_status,
        "durationMs": duration, "created": created,
    })
    return record, created


@transaction.atomic
def submit_voice_transcript(*, actor, record, payload):
    require_reporting_permission(actor, "action.execute", require_shift=True)
    policy = voice_interaction_policy()
    if not policy.enabled or not policy.transcribe_driver_audio:
        raise ReportingError("当前未授权司机语音转文字", "VOICE_TRANSCRIPTION_DISABLED", 409)
    if not isinstance(payload, dict) or set(payload) - VOICE_TRANSCRIPT_KEYS:
        raise ReportingError("转写请求包含未允许字段", "INVALID_VOICE_TRANSCRIPT", 422)
    if str(payload.get("source") or "") != "LOCAL_TRANSCRIBER":
        raise ReportingError("转写结果必须来自受控本地转写适配器", "INVALID_TRANSCRIPTION_SOURCE", 422)
    transcript = _validate_voice_transcript({
        "text": payload.get("transcriptText"), "language": payload.get("language"),
        "utteranceDetected": payload.get("utteranceDetected"), "confidence": payload.get("confidence"),
        "engine": payload.get("engine"),
    }, allow_empty=True)
    record = VoiceInteractionEvidence.objects.select_for_update().select_related("enterprise").get(pk=record.pk)
    if record.requested_by_id != actor.pk:
        raise ReportingError("只能提交本人登记的司机语音转写", "VOICE_EVIDENCE_OWNER_CONFLICT", 403)
    if record.status == VoiceInteractionEvidence.Status.EXPIRED:
        raise ReportingError("语音证据已过期", "VOICE_EVIDENCE_EXPIRED", 410)
    record.transcript = transcript or {}
    record.transcript_engine = (transcript or {}).get("engine", "")
    record.transcript_confidence = (transcript or {}).get("confidence")
    record.transcription_status = VoiceInteractionEvidence.TranscriptionStatus.READY
    record.status = VoiceInteractionEvidence.Status.TRANSCRIBED
    record.save(update_fields=["transcript", "transcript_engine", "transcript_confidence", "transcription_status", "status", "updated_at"])
    audit(actor, "VOICE_TRANSCRIPT_RECORDED", "VOICE_INTERACTION_EVIDENCE", record.public_id, {"engine": record.transcript_engine, "utteranceDetected": (transcript or {}).get("utteranceDetected", False)})
    return record


def get_voice_evidence(*, actor, evidence_id, include_transcript=False):
    require_voice_evidence_permission(actor, include_transcript=include_transcript)
    try:
        parsed_id = uuid.UUID(str(evidence_id))
    except (TypeError, ValueError, AttributeError) as exc:
        raise ReportingError("语音证据标识无效", "INVALID_IDENTIFIER", 400) from exc
    record = VoiceInteractionEvidence.objects.select_related("enterprise").filter(public_id=parsed_id).first()
    if not record:
        raise ReportingError("语音证据不存在", "VOICE_EVIDENCE_NOT_FOUND", 404)
    require_scope(actor, record.enterprise)
    return record, voice_evidence_data(record, include_transcript=include_transcript)


def report_period(period_type, period_value):
    if period_type == ReportSnapshot.PeriodType.DAILY:
        try: start = date.fromisoformat(str(period_value))
        except ValueError as exc: raise ReportingError("日报日期必须为YYYY-MM-DD", "INVALID_REPORT_PERIOD", 422) from exc
        return start, start
    if period_type == ReportSnapshot.PeriodType.MONTHLY:
        try:
            year, month = map(int, str(period_value).split("-"))
            start = date(year, month, 1)
        except (ValueError, TypeError) as exc:
            raise ReportingError("月报月份必须为YYYY-MM", "INVALID_REPORT_PERIOD", 422) from exc
        return start, date(year, month, calendar.monthrange(year, month)[1])
    raise ReportingError("报表类型必须是DAILY或MONTHLY", "INVALID_REPORT_TYPE", 422)


def collapse_same_type_window(facts, window_minutes=None):
    """Keep the last fact in each contiguous same-type reporting window.

    The source facts remain untouched; this is a report aggregation option only.
    """
    if window_minutes in (None, "", 0):
        return facts
    try:
        minutes = int(window_minutes)
    except (TypeError, ValueError) as exc:
        raise ReportingError("同类型报警时间窗口必须是整数分钟", "INVALID_SAME_TYPE_WINDOW", 422) from exc
    if minutes < 1 or minutes > 24 * 60:
        raise ReportingError("同类型报警时间窗口必须介于1至1440分钟", "INVALID_SAME_TYPE_WINDOW", 422)
    grouped = defaultdict(list)
    for fact in facts:
        grouped[(fact.source_kind, fact.alarm_name or "未命名报警")].append(fact)
    selected = []
    window = timedelta(minutes=minutes)
    for rows in grouped.values():
        rows.sort(key=lambda item: item.alarm_time or item.last_seen_at)
        current = None
        current_time = None
        for fact in rows:
            fact_time = fact.alarm_time or fact.last_seen_at
            if current is None or fact_time - current_time <= window:
                current = fact
                current_time = fact_time
            else:
                selected.append(current)
                current = fact
                current_time = fact_time
        if current is not None:
            selected.append(current)
    return sorted(selected, key=lambda item: item.alarm_time or item.last_seen_at)


def metrics_for(enterprise, start, end, cutoff, same_type_window_minutes=None):
    tz = timezone.get_current_timezone()
    start_at = timezone.make_aware(datetime.combine(start, time.min), tz)
    end_at = timezone.make_aware(datetime.combine(end + timedelta(days=1), time.min), tz)
    raw_facts = list(AlarmFact.objects.filter(enterprise=enterprise, alarm_time__gte=start_at, alarm_time__lt=end_at, last_seen_at__lte=cutoff))
    facts = collapse_same_type_window(raw_facts, same_type_window_minutes)
    event_ids = [fact.event_id for fact in facts]
    cases = list(DisposalCase.objects.filter(event_id__in=event_ids))
    source_counts = Counter(fact.source_kind for fact in facts)
    alarm_counts = Counter(fact.alarm_name or "未命名报警" for fact in facts)
    # The customer has confirmed that monthly-report positive reports are
    # exactly server-deduplicated facts from the platform REALTIME source.
    # positiveReportingFlag, appeal status, and action results are downstream
    # disposition fields and must not determine this total.
    # Positive-report totals always count every unique REALTIME AlarmFact in
    # the period. The optional same-type reporting window only affects the
    # general alarm-total presentation and must not erase positive reports.
    positive_facts = [fact for fact in raw_facts if fact.source_kind == "REALTIME"]
    positive_report_type_counts = Counter(fact.alarm_name or "未命名报警" for fact in positive_facts)
    completion_counts = Counter(fact.completion_status or "UNKNOWN_MANUAL" for fact in facts)
    text_attempts = []
    voice_attempts = []
    for fact in facts:
        action = fact.action_snapshot or {}
        attempts = action.get("attempts") if isinstance(action.get("attempts"), list) else []
        if not attempts and action.get("type") == "VOICE_INTERCOM": attempts = [{"channelType": "VOICE", "status": action.get("status")}]
        text_attempts.extend(item for item in attempts if item.get("channelType") == "TEXT")
        voice_attempts.extend(item for item in attempts if item.get("channelType") == "VOICE")
    channel_metrics = lambda attempts: {
        "planned": len(attempts),
        "succeeded": sum(item.get("status") == "SUCCEEDED" for item in attempts),
        "failed": sum(item.get("status") in {"FAILED", "BLOCKED", "UNKNOWN", "MANUAL_REQUIRED"} for item in attempts),
        "unknown": sum(item.get("status") == "UNKNOWN" for item in attempts),
    }
    return {
        "alarmTotal": len(facts), "sourceCounts": dict(sorted(source_counts.items())), "alarmTypeCounts": dict(alarm_counts.most_common()),
        "positiveReportTotal": len(positive_facts),
        "positiveReportTypeCounts": dict(positive_report_type_counts.most_common()),
        "manualCaseTotal": len(cases), "manualPending": sum(case.status != DisposalCase.Status.COMPLETED for case in cases),
        "manualCompleted": sum(case.status == DisposalCase.Status.COMPLETED for case in cases),
        "completion": dict(sorted(completion_counts.items())),
        "text": channel_metrics(text_attempts), "voice": channel_metrics(voice_attempts),
        "dataQuality": {
            "missingAlarmTime": AlarmFact.objects.filter(enterprise=enterprise, alarm_time__isnull=True, first_seen_at__gte=start_at, first_seen_at__lt=end_at, last_seen_at__lte=cutoff).count(),
            "unresolvedCompany": 0,
        },
    }


@transaction.atomic
def generate_report(*, actor, enterprise, period_type, period_value, correction_reason="", same_type_window_minutes=None):
    require_reporting_permission(actor, "report.generate")
    require_scope(actor, enterprise)
    enterprise = EnterpriseScope.objects.select_for_update().get(pk=enterprise.pk)
    start, end = report_period(period_type, period_value)
    cutoff = timezone.now()
    version = (ReportSnapshot.objects.filter(enterprise=enterprise, period_type=period_type, period_start=start).aggregate(Max("version"))["version__max"] or 0) + 1
    correction_reason = str(correction_reason or "").strip()
    if version > 1 and not correction_reason:
        raise ReportingError("更正版报表必须填写更正原因", "CORRECTION_REASON_REQUIRED", 422)
    # Validate the optional aggregation parameter before creating an immutable snapshot.
    if same_type_window_minutes not in (None, "", 0):
        try:
            same_type_window_minutes = int(same_type_window_minutes)
        except (TypeError, ValueError) as exc:
            raise ReportingError("同类型报警时间窗口必须是整数分钟", "INVALID_SAME_TYPE_WINDOW", 422) from exc
        if same_type_window_minutes < 1 or same_type_window_minutes > 24 * 60:
            raise ReportingError("同类型报警时间窗口必须介于1至1440分钟", "INVALID_SAME_TYPE_WINDOW", 422)
    else:
        same_type_window_minutes = None
    snapshot = ReportSnapshot.objects.create(
        enterprise=enterprise, period_type=period_type, period_start=start, period_end=end, version=version,
        metrics=metrics_for(enterprise, start, end, cutoff, same_type_window_minutes),
        parameters={"timezone": "Asia/Shanghai", "periodValue": str(period_value), "sameTypeWindowMinutes": same_type_window_minutes},
        data_cutoff_at=cutoff, generated_by=actor, correction_reason=correction_reason,
    )
    audit(actor, "REPORT_GENERATED", "REPORT_SNAPSHOT", snapshot.public_id, {"enterpriseId": str(enterprise.public_id), "periodType": period_type, "periodStart": start.isoformat(), "version": version})
    return snapshot


@transaction.atomic
def publish_report(*, actor, snapshot):
    require_reporting_permission(actor, "report.publish")
    snapshot = ReportSnapshot.objects.select_for_update().select_related("enterprise").get(pk=snapshot.pk)
    require_scope(actor, snapshot.enterprise)
    if snapshot.status != ReportSnapshot.Status.DRAFT:
        raise ReportingError("只有报表草稿可以发布", "INVALID_REPORT_STATUS", 409)
    for current in ReportSnapshot.objects.select_for_update().filter(enterprise=snapshot.enterprise, period_type=snapshot.period_type, period_start=snapshot.period_start, status=ReportSnapshot.Status.PUBLISHED):
        current.status = ReportSnapshot.Status.RETIRED
        current.save(update_fields=["status", "updated_at"])
    snapshot.status = ReportSnapshot.Status.PUBLISHED
    snapshot.published_by = actor
    snapshot.published_at = timezone.now()
    snapshot.save(update_fields=["status", "published_by", "published_at", "updated_at"])
    audit(actor, "REPORT_PUBLISHED", "REPORT_SNAPSHOT", snapshot.public_id, {"version": snapshot.version})
    return snapshot


def report_rows(snapshot):
    metrics = snapshot.metrics
    rows = [
        ["报表类型", snapshot.get_period_type_display()], ["企业", snapshot.enterprise.name],
        ["统计周期", f"{snapshot.period_start} 至 {snapshot.period_end}"], ["版本", snapshot.version],
        ["数据截止时间", timezone.localtime(snapshot.data_cutoff_at).strftime("%Y-%m-%d %H:%M:%S")],
        ["报警总数", metrics.get("alarmTotal", 0)], ["人工工单", metrics.get("manualCaseTotal", 0)],
        ["正报总数（来源：REALTIME）", metrics.get("positiveReportTotal", 0)],
        ["人工未完成", metrics.get("manualPending", 0)], ["人工已完成", metrics.get("manualCompleted", 0)],
        ["平台已解除", metrics.get("completion", {}).get("PLATFORM_CLEARED", 0)],
        ["平台仍持续", metrics.get("completion", {}).get("PLATFORM_ACTIVE", 0)],
        ["无法判断转人工", metrics.get("completion", {}).get("UNKNOWN_MANUAL", 0)],
        ["文本计划/成功/失败", f"{metrics.get('text', {}).get('planned', 0)}/{metrics.get('text', {}).get('succeeded', 0)}/{metrics.get('text', {}).get('failed', 0)}"],
        ["语音计划/成功/失败", f"{metrics.get('voice', {}).get('planned', 0)}/{metrics.get('voice', {}).get('succeeded', 0)}/{metrics.get('voice', {}).get('failed', 0)}"],
    ]
    rows.append(["报警类型", "数量"])
    rows.extend([[name, count] for name, count in metrics.get("alarmTypeCounts", {}).items()])
    rows.append(["正报类型（均来源：REALTIME）", "数量"])
    rows.extend([[name, count] for name, count in metrics.get("positiveReportTypeCounts", {}).items()])
    return rows


def build_xlsx(snapshot):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "日报" if snapshot.period_type == "DAILY" else "月报"
    for row in report_rows(snapshot): sheet.append(row)
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 42
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="175F47")
    for row in sheet.iter_rows():
        for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_pdf(snapshot):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.pdfgen import canvas
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    output = BytesIO()
    canvas_obj = canvas.Canvas(output, pagesize=A4)
    width, height = A4
    canvas_obj.setFont("STSong-Light", 18)
    canvas_obj.drawString(48, height - 55, f"{snapshot.enterprise.name} {snapshot.get_period_type_display()}")
    y = height - 90
    canvas_obj.setFont("STSong-Light", 10)
    for label, value in report_rows(snapshot):
        if y < 50:
            canvas_obj.showPage(); canvas_obj.setFont("STSong-Light", 10); y = height - 50
        canvas_obj.drawString(48, y, str(label)[:30])
        canvas_obj.drawString(220, y, str(value)[:55])
        y -= 18
    canvas_obj.save()
    return output.getvalue()


@transaction.atomic
def create_export(*, actor, snapshot, format_name, purpose):
    require_reporting_permission(actor, "export.masked")
    snapshot = ReportSnapshot.objects.select_related("enterprise").get(pk=snapshot.pk)
    require_scope(actor, snapshot.enterprise)
    if snapshot.status != ReportSnapshot.Status.PUBLISHED:
        raise ReportingError("只有已发布报表可以导出", "INVALID_REPORT_STATUS", 409)
    purpose = str(purpose or "").strip()
    if not purpose or len(purpose) > 500:
        raise ReportingError("导出用途不能为空且不能超过500字符", "INVALID_EXPORT_PURPOSE", 422)
    if format_name != ExportJob.Format.XLSX:
        raise ReportingError("当前阶段只支持Excel XLSX导出", "INVALID_EXPORT_FORMAT", 422)
    content, suffix = build_xlsx(snapshot), "xlsx"
    export_dir = Path(settings.REPORT_EXPORT_DIR).resolve()
    export_dir.mkdir(parents=True, exist_ok=True)
    file_name = f"{snapshot.enterprise.code}-{snapshot.period_type.lower()}-{snapshot.period_start}-v{snapshot.version}.{suffix}"
    path = (export_dir / f"{uuid_safe()}-{file_name}").resolve()
    if export_dir not in path.parents: raise ReportingError("导出路径无效", "INVALID_EXPORT_PATH", 500)
    path.write_bytes(content)
    job = ExportJob.objects.create(
        report_snapshot=snapshot, format=format_name, purpose=purpose, file_name=file_name, file_path=str(path),
        file_sha256=hashlib.sha256(content).hexdigest(), file_size=len(content), created_by=actor,
        expires_at=timezone.now() + timedelta(days=7),
    )
    audit(actor, "REPORT_EXPORTED", "EXPORT_JOB", job.public_id, {"snapshotId": str(snapshot.public_id), "format": format_name, "fileHash": job.file_sha256, "purpose": purpose})
    return job


def uuid_safe():
    import uuid
    return uuid.uuid4().hex


@transaction.atomic
def record_download(*, actor, job):
    require_reporting_permission(actor, "export.masked")
    job = ExportJob.objects.select_for_update().select_related("report_snapshot__enterprise").get(pk=job.pk)
    require_scope(actor, job.report_snapshot.enterprise)
    if job.status != ExportJob.Status.READY or job.expires_at <= timezone.now():
        if job.status == ExportJob.Status.READY:
            job.status = ExportJob.Status.EXPIRED
            job.save(update_fields=["status", "updated_at"])
        raise ReportingError("导出文件已过期或不可下载", "EXPORT_NOT_AVAILABLE", 410)
    path = Path(job.file_path).resolve()
    export_dir = Path(settings.REPORT_EXPORT_DIR).resolve()
    if export_dir not in path.parents or not path.is_file():
        raise ReportingError("导出文件不存在", "EXPORT_FILE_MISSING", 404)
    content = path.read_bytes()
    if hashlib.sha256(content).hexdigest() != job.file_sha256:
        raise ReportingError("导出文件完整性校验失败", "EXPORT_HASH_MISMATCH", 409)
    job.download_count += 1
    job.last_downloaded_at = timezone.now()
    job.save(update_fields=["download_count", "last_downloaded_at", "updated_at"])
    audit(actor, "REPORT_DOWNLOADED", "EXPORT_JOB", job.public_id, {"downloadCount": job.download_count, "fileHash": job.file_sha256})
    return job, content
