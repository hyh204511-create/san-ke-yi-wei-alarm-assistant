import hashlib
import json
import tempfile
from pathlib import Path
from datetime import timedelta

from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse
from django.core.management import call_command
from django.utils import timezone
from django.db import connection
from openpyxl import load_workbook

from apps.disposals.models import DisposalCase
from apps.governance.models import (
    AssistantProfile, AuditEvent, DeviceRegistration, EnterpriseGrant, EnterpriseScope,
    RoleAssignment, VoiceInteractionPolicy,
)
from apps.governance.services import assign_role, claim_shift

from . import services
from .models import ActionLease, AlarmFact, CaptureSource, DutyNotification, ExportJob, ReportSnapshot, VoiceInteractionEvidence


def event_payload(event_id="alarm:id:9000000000000000001", *, source_kind="REALTIME", alarm_name="疲劳驾驶报警"):
    return {
        "eventId": event_id, "alarmId": event_id.split(":")[-1], "sourceKind": source_kind, "alarmName": alarm_name,
        "alarmTime": "2026-07-19 12:10:00", "discoveredAt": "2026-07-19T12:10:05+08:00", "updatedAt": "2026-07-19T12:10:06+08:00",
        "vehicleId": "vehicle-001", "vehicleNo": "湘A测001", "companyId": "REPORT-COMPANY", "companyName": "报表测试企业", "state": "MANUAL_REQUIRED",
        "completionAssessment": {"status": "UNKNOWN_MANUAL", "source": "PLATFORM_STATUS", "manualRequired": True, "reason": "测试状态字段缺失"},
    }


def decision_payload():
    return {"decisionId": "decision-report-1", "action": "RESPONSE_PLAN", "ruleSetVersion": "rules-v2", "channels": [{"type": "TEXT"}, {"type": "VOICE"}]}


def action_payload():
    return {"actionId": "plan-report-1", "type": "RESPONSE_PLAN", "status": "MANUAL_REQUIRED", "attempts": [
        {"channelType": "TEXT", "status": "SUCCEEDED"}, {"channelType": "VOICE", "status": "FAILED"},
    ]}


class ReportingFlowTests(TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.settings_override = override_settings(REPORT_EXPORT_DIR=self.temp_dir.name)
        self.settings_override.enable()
        users = get_user_model()
        self.monitor = users.objects.create_user(username="report-monitor")
        self.reporter = users.objects.create_user(username="reporter")
        self.outsider = users.objects.create_user(username="report-outsider")
        self.reviewer = users.objects.create_user(username="evidence-reviewer")
        for user, name, code in [
            (self.monitor, "报表监控员", "RPT-MON"), (self.reporter, "数据报表员", "RPT-USER"),
            (self.outsider, "外部报表员", "RPT-OUT"), (self.reviewer, "证据审核员", "RPT-REVIEW"),
        ]:
            AssistantProfile.objects.create(user=user, display_name=name, employee_code=code)
        self.enterprise = EnterpriseScope.objects.create(code="REPORT-COMPANY", name="报表测试企业", scope_type=EnterpriseScope.ScopeType.ENTERPRISE)
        outside = EnterpriseScope.objects.create(code="REPORT-OUT", name="外部企业", scope_type=EnterpriseScope.ScopeType.ENTERPRISE)
        EnterpriseGrant.objects.create(user=self.monitor, enterprise=self.enterprise)
        EnterpriseGrant.objects.create(user=self.reporter, enterprise=self.enterprise)
        EnterpriseGrant.objects.create(user=self.outsider, enterprise=outside)
        EnterpriseGrant.objects.create(user=self.reviewer, enterprise=self.enterprise)
        assign_role(user=self.monitor, role=RoleAssignment.Role.MONITOR_OPERATOR, assigned_by=self.monitor)
        assign_role(user=self.reporter, role=RoleAssignment.Role.SYSTEM_ADMIN, assigned_by=self.reporter)
        assign_role(user=self.outsider, role=RoleAssignment.Role.UNIT_USER, assigned_by=self.outsider)
        assign_role(user=self.reviewer, role=RoleAssignment.Role.RULE_REVIEWER, assigned_by=self.reviewer)
        claim_shift(user=self.monitor, platform_account_ref="report-platform", workstation_id="REPORT-WS")
        claim_shift(user=self.outsider, platform_account_ref="report-platform-outside", workstation_id="REPORT-WS-OUTSIDE")

    def tearDown(self):
        self.settings_override.disable()
        self.temp_dir.cleanup()

    def post_json(self, url, data, *, action_token=False):
        headers = {}
        if action_token:
            token = self.client.get(reverse("assistant-action-token-api")).json()["data"]["actionToken"]
            headers["HTTP_X_ASSISTANT_ACTION_TOKEN"] = token
        return self.client.post(url, data=json.dumps(data), content_type="application/json", **headers)

    def ingest(self):
        self.client.force_login(self.monitor)
        response = self.post_json(reverse("report-event-upsert-api"), {"event": event_payload(), "decision": decision_payload(), "action": action_payload()}, action_token=True)
        self.assertEqual(response.status_code, 201)
        return response

    def generate_and_publish(self, period_type="DAILY", period_value="2026-07-19"):
        self.client.force_login(self.reporter)
        generated = self.post_json(reverse("report-generate-api"), {
            "enterpriseId": str(self.enterprise.public_id), "periodType": period_type, "periodValue": period_value,
        })
        self.assertEqual(generated.status_code, 201)
        report = generated.json()["data"]
        published = self.post_json(reverse("report-publish-api", args=[report["reportId"]]), {})
        self.assertEqual(published.status_code, 200)
        return published.json()["data"]

    def completed_voice_lease(self):
        self.ingest()
        fact = AlarmFact.objects.get()
        self.assertEqual(fact.completion_status, "UNKNOWN_MANUAL")
        self.assertTrue(fact.completion_manual_required)
        lease = services.acquire_action_lease(
            actor=self.monitor, fact=fact, device_id="voice-device", action_type="VOICE_INTERCOM", mode="SANDBOX",
        )
        lease.status = ActionLease.Status.COMPLETED
        lease.result_code = "SUCCEEDED"
        lease.finished_at = timezone.now()
        lease.save(update_fields=["status", "result_code", "finished_at", "updated_at"])
        return fact, lease

    def enable_voice_evidence(self, *, transcribe=True, retention_days=7):
        return VoiceInteractionPolicy.objects.create(
            enabled=True, record_driver_audio=True, transcribe_driver_audio=transcribe,
            retention_days=retention_days,
        )

    def test_daily_report_uses_immutable_cutoff_and_channel_metrics(self):
        self.ingest()
        DisposalCase.objects.create(
            event_id=event_payload()["eventId"], alarm_id=event_payload()["alarmId"], enterprise=self.enterprise, source_kind="REALTIME",
            alarm_name="疲劳驾驶报警", vehicle_no="湘A测001", event_snapshot=event_payload(), latest_event_snapshot=event_payload(),
            decision_snapshot={"action": "MANUAL_REVIEW"}, requires_review=True,
        )
        report = self.generate_and_publish()
        self.assertEqual(report["metrics"]["alarmTotal"], 1)
        self.assertEqual(report["metrics"]["manualPending"], 1)
        self.assertEqual(report["metrics"]["completion"]["UNKNOWN_MANUAL"], 1)
        self.assertEqual(report["metrics"]["text"]["succeeded"], 1)
        self.assertEqual(report["metrics"]["voice"]["failed"], 1)
        self.assertEqual(report["parameters"]["timezone"], "Asia/Shanghai")

    def test_monthly_positive_reports_are_realtime_only_and_grouped_by_type(self):
        self.ingest()
        # PREWARNING remains part of the alarm report, but is not a positive
        # report under the customer-confirmed monthly-report definition.
        prewarning = event_payload(
            "alarm:id:9000000000000000002", source_kind="PREWARNING", alarm_name="超速预警",
        )
        self.client.force_login(self.monitor)
        response = self.post_json(
            reverse("report-event-upsert-api"),
            {"event": prewarning, "decision": decision_payload(), "action": {}},
            action_token=True,
        )
        self.assertEqual(response.status_code, 201)

        realtime_second = event_payload("alarm:id:9000000000000000003", alarm_name="抽烟报警")
        response = self.post_json(
            reverse("report-event-upsert-api"),
            {"event": realtime_second, "decision": decision_payload(), "action": {}},
            action_token=True,
        )
        self.assertEqual(response.status_code, 201)

        report = self.generate_and_publish("MONTHLY", "2026-07")
        metrics = report["metrics"]
        self.assertEqual(metrics["alarmTotal"], 3)
        self.assertEqual(metrics["positiveReportTotal"], 2)
        self.assertEqual(metrics["positiveReportTypeCounts"], {event_payload()["alarmName"]: 1, "抽烟报警": 1})
        self.assertNotIn("超速预警", metrics["positiveReportTypeCounts"])

    def test_sensitive_alarm_snapshots_are_encrypted_at_rest(self):
        self.ingest()
        fact = AlarmFact.objects.get()
        self.assertEqual(fact.event_snapshot["vehicleNo"], "湘A测001")
        with connection.cursor() as cursor:
            cursor.execute("SELECT event_snapshot, decision_snapshot, action_snapshot, ingestion_provenance FROM assistant_alarm_facts WHERE id = %s", [fact.pk])
            raw = cursor.fetchone()
        self.assertTrue(all(str(value).startswith("enc:v1:") for value in raw))
        self.assertNotIn("湘A测001", raw[0])

    def test_overlapping_accounts_upsert_one_event_and_preserve_sources(self):
        self.ingest()
        second = get_user_model().objects.create_user(username="report-second-collector")
        AssistantProfile.objects.create(user=second, display_name="第二采集人员", employee_code="RPT-SECOND")
        EnterpriseGrant.objects.create(user=second, enterprise=self.enterprise)
        assign_role(user=second, role=RoleAssignment.Role.MONITOR_OPERATOR, assigned_by=second)
        claim_shift(user=second, platform_account_ref="report-platform-second", workstation_id="REPORT-WS-SECOND")
        self.client.force_login(second)
        response = self.post_json(
            reverse("report-event-upsert-api"),
            {"event": event_payload(), "decision": decision_payload(), "action": action_payload()},
            action_token=True,
        )
        self.assertIn(response.status_code, {200, 201})
        self.assertEqual(AlarmFact.objects.count(), 1)
        fact = AlarmFact.objects.get()
        self.assertEqual(fact.ingested_by, self.monitor)
        self.assertEqual(len(fact.ingestion_provenance), 2)
        self.assertEqual({row["workstationId"] for row in fact.ingestion_provenance}, {"REPORT-WS", "REPORT-WS-SECOND"})

    def test_fifty_idempotent_capture_sources_keep_one_fact(self):
        self.client.force_login(self.monitor)
        for index in range(50):
            response = self.post_json(
                reverse("report-event-upsert-api"),
                {"event": event_payload(), "decision": decision_payload(), "action": action_payload(), "source": {
                    "captureId": f"capture-{index}", "deviceId": f"device-{index}",
                    "platformAccountRef": "report-platform", "extensionVersion": "0.4.0",
                    "endpoint": "/api/alarm-service/alarm/center/query", "capturedAt": "2026-07-19T12:10:06+08:00",
                }}, action_token=True,
            )
            self.assertIn(response.status_code, {200, 201})
        self.assertEqual(AlarmFact.objects.count(), 1)
        self.assertEqual(CaptureSource.objects.count(), 50)

    def test_only_one_active_plan_lease_per_alarm_across_all_channels(self):
        self.ingest()
        fact = AlarmFact.objects.get()
        first = services.acquire_action_lease(actor=self.monitor, fact=fact, device_id="device-one", action_type="RESPONSE_PLAN")
        self.assertEqual(first.status, ActionLease.Status.ACTIVE)
        with self.assertRaises(services.ReportingError) as caught:
            services.acquire_action_lease(actor=self.monitor, fact=fact, device_id="device-two", action_type="VOICE")
        self.assertEqual(caught.exception.code, "ACTION_LEASE_CONFLICT")
        first.status = ActionLease.Status.EXECUTING
        first.save(update_fields=["status", "updated_at"])
        with self.assertRaises(services.ReportingError) as caught_executing:
            services.acquire_action_lease(actor=self.monitor, fact=fact, device_id="device-three", action_type="TEXT")
        self.assertEqual(caught_executing.exception.code, "ACTION_LEASE_CONFLICT")

    def test_completed_plan_and_unknown_action_both_block_automatic_replay(self):
        self.ingest()
        fact = AlarmFact.objects.get()
        completed = services.acquire_action_lease(
            actor=self.monitor, fact=fact, device_id="device-completed", action_type="PLAN", mode="SANDBOX",
        )
        completed.status = ActionLease.Status.COMPLETED
        completed.result_code = "SUCCEEDED"
        completed.finished_at = timezone.now()
        completed.save(update_fields=["status", "result_code", "finished_at", "updated_at"])
        with self.assertRaises(services.ReportingError) as already_completed:
            services.acquire_action_lease(
                actor=self.monitor, fact=fact, device_id="device-replay", action_type="RESPONSE_PLAN", mode="SANDBOX",
            )
        self.assertEqual(already_completed.exception.code, "ACTION_ALREADY_COMPLETED")

        completed.action_type = "TEXT_TTS"
        completed.status = ActionLease.Status.UNKNOWN
        completed.result_code = "UNKNOWN"
        completed.save(update_fields=["action_type", "status", "result_code", "updated_at"])
        with self.assertRaises(services.ReportingError) as unknown:
            services.acquire_action_lease(
                actor=self.monitor, fact=fact, device_id="device-unknown-replay", action_type="RESPONSE_PLAN", mode="SANDBOX",
            )
        self.assertEqual(unknown.exception.code, "ACTION_RESULT_UNKNOWN_MANUAL")

    def test_processing_status_tracks_plan_result_and_survives_empty_refresh(self):
        self.client.force_login(self.monitor)
        event = event_payload("alarm:id:9000000000000000088", alarm_name="超速驾驶")
        created = self.post_json(
            reverse("report-event-upsert-api"),
            {"event": event, "decision": decision_payload(), "action": {}}, action_token=True,
        )
        self.assertEqual(created.status_code, 201)
        fact = AlarmFact.objects.get(event_id=event["eventId"])
        self.assertEqual(fact.processing_status, AlarmFact.ProcessingStatus.UNPROCESSED)

        lease = services.acquire_action_lease(
            actor=self.monitor, fact=fact, device_id="processing-device", action_type="RESPONSE_PLAN", mode="SANDBOX",
        )
        fact.refresh_from_db()
        self.assertEqual(fact.processing_status, AlarmFact.ProcessingStatus.EXECUTING)
        services.record_action_result(actor=self.monitor, payload={
            "leaseId": str(lease.public_id), "leaseToken": lease._plain_token,
            "deviceId": "processing-device", "resultCode": "SUCCEEDED",
            "actionId": "processing-plan", "result": {"processingStatus": "PROCESSED"},
        })
        fact.refresh_from_db()
        self.assertEqual(fact.processing_status, AlarmFact.ProcessingStatus.PROCESSED)
        self.assertEqual(fact.processing_source, "SERVER_ACTION_RESULT")
        self.assertIsNotNone(fact.processing_marked_at)

        stale_action = {"actionId": "stale-plan", "type": "RESPONSE_PLAN", "status": "EXECUTING", "processingStatus": "EXECUTING"}
        refreshed = self.post_json(
            reverse("report-event-upsert-api"),
            {"event": event, "decision": decision_payload(), "action": stale_action}, action_token=True,
        )
        self.assertEqual(refreshed.status_code, 200)
        fact.refresh_from_db()
        self.assertEqual(fact.processing_status, AlarmFact.ProcessingStatus.PROCESSED)
        ActionLease.objects.filter(fact=fact).delete()
        with self.assertRaises(services.ReportingError) as replay:
            services.acquire_action_lease(
                actor=self.monitor, fact=fact, device_id="processing-replay", action_type="RESPONSE_PLAN", mode="SANDBOX",
            )
        self.assertEqual(replay.exception.code, "ACTION_ALREADY_COMPLETED")

    def test_unknown_plan_result_is_explicit_and_blocks_cross_account_replay(self):
        self.client.force_login(self.monitor)
        event = event_payload("alarm:id:9000000000000000089", alarm_name="超速驾驶")
        self.post_json(
            reverse("report-event-upsert-api"),
            {"event": event, "decision": decision_payload(), "action": {}}, action_token=True,
        )
        fact = AlarmFact.objects.get(event_id=event["eventId"])
        lease = services.acquire_action_lease(
            actor=self.monitor, fact=fact, device_id="unknown-device", action_type="RESPONSE_PLAN", mode="SANDBOX",
        )
        services.record_action_result(actor=self.monitor, payload={
            "leaseId": str(lease.public_id), "leaseToken": lease._plain_token,
            "deviceId": "unknown-device", "resultCode": "UNKNOWN",
            "actionId": "unknown-plan", "result": {"errorCode": "PLATFORM_REQUEST_TIMEOUT"},
        })
        fact.refresh_from_db()
        self.assertEqual(fact.processing_status, AlarmFact.ProcessingStatus.UNKNOWN)
        DutyNotification.objects.filter(action_lease=lease).delete()
        fact.action_leases.all().delete()
        with self.assertRaises(services.ReportingError) as caught:
            services.acquire_action_lease(
                actor=self.monitor, fact=fact, device_id="second-device", action_type="RESPONSE_PLAN", mode="SANDBOX",
            )
        self.assertEqual(caught.exception.code, "ACTION_RESULT_UNKNOWN_MANUAL")

    def test_expired_or_manual_action_state_fails_closed_before_new_lease(self):
        self.client.force_login(self.monitor)
        event = event_payload("alarm:id:9000000000000000090", alarm_name="超速驾驶")
        self.post_json(
            reverse("report-event-upsert-api"),
            {"event": event, "decision": decision_payload(), "action": {}}, action_token=True,
        )
        fact = AlarmFact.objects.get(event_id=event["eventId"])
        lease = services.acquire_action_lease(
            actor=self.monitor, fact=fact, device_id="expired-device", action_type="RESPONSE_PLAN", mode="SANDBOX",
        )
        lease.expires_at = timezone.now() - timedelta(seconds=1)
        lease.save(update_fields=["expires_at", "updated_at"])
        with self.assertRaises(services.ReportingError) as expired:
            services.acquire_action_lease(
                actor=self.monitor, fact=fact, device_id="second-device", action_type="RESPONSE_PLAN", mode="SANDBOX",
            )
        self.assertEqual(expired.exception.code, "ACTION_RESULT_UNKNOWN_MANUAL")

        call_command("expire_action_leases", verbosity=0)
        fact.refresh_from_db()
        self.assertEqual(fact.processing_status, AlarmFact.ProcessingStatus.UNKNOWN)

        manual_event = event_payload("alarm:id:9000000000000000091", alarm_name="超速驾驶")
        self.post_json(
            reverse("report-event-upsert-api"),
            {"event": manual_event, "decision": decision_payload(), "action": {}}, action_token=True,
        )
        manual_fact = AlarmFact.objects.get(event_id=manual_event["eventId"])
        manual_lease = services.acquire_action_lease(
            actor=self.monitor, fact=manual_fact, device_id="manual-device", action_type="RESPONSE_PLAN", mode="SANDBOX",
        )
        services.record_action_result(actor=self.monitor, payload={
            "leaseId": str(manual_lease.public_id), "leaseToken": manual_lease._plain_token,
            "deviceId": "manual-device", "resultCode": "FAILED", "actionId": "manual-plan", "result": {},
        })
        DutyNotification.objects.filter(action_lease=manual_lease).delete()
        manual_fact.action_leases.all().delete()
        with self.assertRaises(services.ReportingError) as manual:
            services.acquire_action_lease(
                actor=self.monitor, fact=manual_fact, device_id="manual-replay", action_type="RESPONSE_PLAN", mode="SANDBOX",
            )
        self.assertEqual(manual.exception.code, "ACTION_MANUAL_REVIEW_REQUIRED")

    def test_action_lease_failure_creates_duty_notification_and_can_be_acknowledged(self):
        self.ingest()
        DutyNotification.objects.all().delete()
        fact = AlarmFact.objects.get()
        DeviceRegistration.objects.create(
            device_id="device-action-1", user=self.monitor, platform_account_ref="report-platform",
            extension_version="0.5.3", session_status="AUTHENTICATED", platform_identity_status="UNKNOWN",
        )
        self.client.force_login(self.monitor)
        acquire = self.post_json(reverse("report-action-lease-acquire-api"), {
            "eventId": fact.event_id, "deviceId": "device-action-1", "actionType": "TEXT_TTS",
            "mode": "SANDBOX", "durationSeconds": 90,
        }, action_token=True)
        self.assertEqual(acquire.status_code, 201)
        lease_data = acquire.json()["data"]
        self.assertTrue(lease_data["leaseToken"])
        result = self.post_json(reverse("report-action-lease-result-api", args=[lease_data["leaseId"]]), {
            "leaseToken": lease_data["leaseToken"], "deviceId": "device-action-1", "resultCode": "FAILED",
            "actionId": "action-report-failure", "result": {"errorCode": "TEXT_FAILED", "simulated": True},
        }, action_token=True)
        self.assertEqual(result.status_code, 200)
        self.assertTrue(result.json()["data"]["notificationCreated"])
        self.assertEqual(ActionLease.objects.get(public_id=lease_data["leaseId"]).status, ActionLease.Status.FAILED)
        self.assertEqual(DutyNotification.objects.filter(recipient=self.monitor).count(), 1)
        notifications = self.client.get(reverse("report-notifications-api"))
        self.assertEqual(notifications.status_code, 200)
        notification_id = notifications.json()["data"][0]["notificationId"]
        ack = self.post_json(reverse("report-notification-ack-api", args=[notification_id]), {}, action_token=True)
        self.assertEqual(ack.status_code, 200)
        self.assertEqual(DutyNotification.objects.get().status, DutyNotification.Status.ACKNOWLEDGED)

    def test_successful_voice_action_creates_review_notification(self):
        self.ingest()
        DutyNotification.objects.all().delete()
        fact = AlarmFact.objects.get()
        lease = services.acquire_action_lease(
            actor=self.monitor, fact=fact, device_id="voice-review-device", action_type="VOICE_INTERCOM", mode="SANDBOX",
        )
        lease, notified = services.record_action_result(
            actor=self.monitor,
            payload={
                "leaseId": str(lease.public_id), "leaseToken": lease._plain_token, "deviceId": "voice-review-device",
                "resultCode": "SUCCEEDED", "actionId": "voice-review-action", "result": {"simulated": True, "playbackStarted": True},
            },
        )
        self.assertTrue(notified)
        self.assertEqual(lease.status, ActionLease.Status.COMPLETED)
        notice = DutyNotification.objects.get(recipient=self.monitor)
        self.assertEqual(notice.kind, "VOICE_REVIEW_REQUIRED")
        self.assertEqual(notice.result_code, "SUCCEEDED")

    def test_live_action_lease_requires_verified_platform_identity(self):
        self.ingest()
        fact = AlarmFact.objects.get()
        DeviceRegistration.objects.create(
            device_id="device-action-unknown", user=self.monitor, platform_account_ref="report-platform",
            extension_version="0.5.3", session_status="AUTHENTICATED", platform_identity_status="UNKNOWN",
        )
        self.client.force_login(self.monitor)
        response = self.post_json(reverse("report-action-lease-acquire-api"), {
            "eventId": fact.event_id, "deviceId": "device-action-unknown", "actionType": "VOICE_INTERCOM", "mode": "LIVE",
        }, action_token=True)
        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()["code"], "PLATFORM_IDENTITY_REQUIRED")

    def test_voice_evidence_is_rejected_when_recording_policy_is_closed(self):
        _fact, lease = self.completed_voice_lease()
        with self.assertRaises(services.ReportingError) as caught:
            services.register_voice_evidence(
                actor=self.monitor,
                lease_id=lease.public_id,
                payload={"leaseToken": lease._plain_token, "deviceId": "voice-device"},
            )
        self.assertEqual(caught.exception.code, "VOICE_RECORDING_DISABLED")

    def test_voice_evidence_requires_consent_and_local_transcriber(self):
        _fact, lease = self.completed_voice_lease()
        self.enable_voice_evidence()
        base = {"leaseToken": lease._plain_token, "deviceId": "voice-device", "source": "LOCAL_AUDIO_CAPTURE"}
        with self.assertRaises(services.ReportingError) as missing_consent:
            services.register_voice_evidence(actor=self.monitor, lease_id=lease.public_id, payload=base)
        self.assertEqual(missing_consent.exception.code, "VOICE_CONSENT_REQUIRED")
        with self.assertRaises(services.ReportingError) as missing_reference:
            services.register_voice_evidence(
                actor=self.monitor, lease_id=lease.public_id,
                payload={**base, "consentAccepted": True},
            )
        self.assertEqual(missing_reference.exception.code, "VOICE_CONSENT_REFERENCE_REQUIRED")
        with self.assertRaises(services.ReportingError) as remote_transcript:
            services.register_voice_evidence(
                actor=self.monitor, lease_id=lease.public_id,
                payload={
                    **base, "consentAccepted": True, "consentReference": "consent-1",
                    "transcript": {"text": "远程转写", "language": "zh-CN", "utteranceDetected": True, "engine": "remote"},
                    "source": "REMOTE_TRANSCRIBER",
                },
            )
        self.assertEqual(remote_transcript.exception.code, "INVALID_TRANSCRIPTION_SOURCE")

    def test_voice_evidence_transcript_is_encrypted_and_review_permission_is_separate(self):
        _fact, lease = self.completed_voice_lease()
        self.enable_voice_evidence()
        record, created = services.register_voice_evidence(
            actor=self.monitor, lease_id=lease.public_id,
            payload={
                "leaseToken": lease._plain_token, "deviceId": "voice-device", "audioSha256": "a" * 64,
                "durationMs": 1200, "audioRef": "evidence/voice-001",
                "recordedStartedAt": "2026-07-19T12:11:00+08:00", "recordedEndedAt": "2026-07-19T12:11:01+08:00",
                "source": "LOCAL_TRANSCRIBER", "consentAccepted": True, "consentReference": "consent-1",
                "transcript": {"text": "请注意安全驾驶", "language": "zh-CN", "utteranceDetected": True, "confidence": 0.91, "engine": "local-v1"},
            },
        )
        self.assertTrue(created)
        self.assertEqual(record.status, VoiceInteractionEvidence.Status.TRANSCRIBED)
        with connection.cursor() as cursor:
            cursor.execute("SELECT audio_metadata, transcript FROM assistant_voice_interaction_evidence WHERE id = %s", [record.pk])
            raw_audio_metadata, raw_transcript = cursor.fetchone()
        self.assertTrue(str(raw_audio_metadata).startswith("enc:v1:"))
        self.assertTrue(str(raw_transcript).startswith("enc:v1:"))
        self.assertNotIn("请注意安全驾驶", str(raw_transcript))

        self.client.force_login(self.monitor)
        metadata = self.client.get(reverse("report-voice-evidence-detail-api", args=[record.public_id]))
        self.assertEqual(metadata.status_code, 200)
        self.assertNotIn("transcript", metadata.json()["data"])
        denied = self.client.get(reverse("report-voice-evidence-detail-api", args=[record.public_id]) + "?includeTranscript=1")
        self.assertEqual(denied.status_code, 403)

        self.client.force_login(self.reviewer)
        reviewed = self.client.get(reverse("report-voice-evidence-detail-api", args=[record.public_id]) + "?includeTranscript=1")
        self.assertEqual(reviewed.status_code, 200)
        self.assertEqual(reviewed.json()["data"]["transcript"]["text"], "请注意安全驾驶")

    def test_voice_evidence_expiry_clears_audio_reference_and_transcript(self):
        _fact, lease = self.completed_voice_lease()
        self.enable_voice_evidence(retention_days=1)
        record, _created = services.register_voice_evidence(
            actor=self.monitor, lease_id=lease.public_id,
            payload={
                "leaseToken": lease._plain_token, "deviceId": "voice-device", "durationMs": 100,
                "audioRef": "evidence/voice-expiring", "source": "LOCAL_TRANSCRIBER",
                "consentAccepted": True, "consentReference": "consent-expiring",
                "transcript": {"text": "司机未回复", "language": "zh-CN", "utteranceDetected": True, "engine": "local-v1"},
            },
        )
        record.retention_until = timezone.now() - timedelta(seconds=1)
        record.save(update_fields=["retention_until", "updated_at"])
        call_command("purge_expired_voice_evidence", dry_run=True, verbosity=0)
        call_command("purge_expired_voice_evidence", verbosity=0)
        record.refresh_from_db()
        self.assertEqual(record.status, VoiceInteractionEvidence.Status.EXPIRED)
        self.assertEqual(record.transcript, {})
        self.assertEqual(record.audio_metadata, {"purged": True})

    def test_expire_action_leases_marks_unknown_and_notifies(self):
        self.ingest()
        DutyNotification.objects.all().delete()
        fact = AlarmFact.objects.get()
        lease = services.acquire_action_lease(actor=self.monitor, fact=fact, device_id="device-timeout", action_type="VOICE_INTERCOM")
        lease.expires_at = timezone.now() - timedelta(seconds=1)
        lease.save(update_fields=["expires_at", "updated_at"])
        call_command("expire_action_leases", verbosity=0)
        lease.refresh_from_db()
        fact.refresh_from_db()
        self.assertEqual(lease.status, ActionLease.Status.UNKNOWN)
        self.assertEqual(fact.processing_status, AlarmFact.ProcessingStatus.UNKNOWN)
        self.assertEqual(fact.processing_source, "SERVER_LEASE_TIMEOUT")
        self.assertEqual(DutyNotification.objects.filter(recipient=self.monitor, result_code="UNKNOWN").count(), 1)

    def test_monthly_report_and_correction_create_new_versions(self):
        self.ingest()
        first = self.generate_and_publish("MONTHLY", "2026-07")
        corrected = self.post_json(reverse("report-generate-api"), {
            "enterpriseId": str(self.enterprise.public_id), "periodType": "MONTHLY", "periodValue": "2026-07", "correctionReason": "补录迟到报警",
        })
        self.assertEqual(corrected.status_code, 201)
        self.assertEqual(corrected.json()["data"]["version"], 2)
        self.assertEqual(first["version"], 1)

    def test_report_can_keep_last_same_type_fact_in_optional_window(self):
        self.ingest()
        second = event_payload("alarm:id:9000000000000000002")
        second["alarmTime"] = "2026-07-19 12:20:00"
        second["discoveredAt"] = "2026-07-19T12:20:05+08:00"
        second["updatedAt"] = "2026-07-19T12:20:06+08:00"
        self.client.force_login(self.monitor)
        response = self.post_json(reverse("report-event-upsert-api"), {
            "event": second, "decision": decision_payload(), "action": {"actionId": "action-2", "status": "SUCCEEDED"},
            "source": {"captureId": "capture-2", "deviceId": "REPORT-WS", "platformAccountRef": "report-platform", "capturedAt": second["updatedAt"]},
        }, action_token=True)
        self.assertIn(response.status_code, {200, 201})
        report = self.generate_and_publish()
        # The two same-type events are within a 30-minute aggregation window.
        self.client.force_login(self.reporter)
        generated = self.post_json(reverse("report-generate-api"), {
            "enterpriseId": str(self.enterprise.public_id), "periodType": "DAILY", "periodValue": "2026-07-19",
            "sameTypeWindowMinutes": 30, "correctionReason": "确认统计口径",
        })
        self.assertEqual(generated.status_code, 201)
        self.assertEqual(generated.json()["data"]["metrics"]["alarmTotal"], 1)
        self.assertEqual(generated.json()["data"]["metrics"]["positiveReportTotal"], 2)
        self.assertEqual(generated.json()["data"]["parameters"]["sameTypeWindowMinutes"], 30)

    def test_correction_requires_reason(self):
        self.ingest()
        self.generate_and_publish()
        response = self.post_json(reverse("report-generate-api"), {
            "enterpriseId": str(self.enterprise.public_id), "periodType": "DAILY", "periodValue": "2026-07-19",
        })
        self.assertEqual(response.status_code, 422)
        self.assertEqual(response.json()["code"], "CORRECTION_REASON_REQUIRED")

    def test_xlsx_exports_are_hashed_audited_and_downloadable(self):
        self.ingest()
        report = self.generate_and_publish()
        response = self.post_json(reverse("report-export-api", args=[report["reportId"]]), {"format": "XLSX", "purpose": "企业月度安全例会"})
        self.assertEqual(response.status_code, 201)
        export = response.json()["data"]
        download = self.client.get(reverse("report-download-api", args=[export["exportId"]]))
        content = b"".join(download.streaming_content)
        self.assertTrue(content.startswith(b"PK"))
        self.assertEqual(hashlib.sha256(content).hexdigest(), export["fileHash"])
        workbook = load_workbook(filename=Path(ExportJob.objects.get(public_id=export["exportId"]).file_path), read_only=True)
        self.assertIn("日报", workbook.sheetnames)
        rows = list(workbook["日报"].iter_rows(values_only=True))
        self.assertTrue(any(row[0] == "正报总数（来源：REALTIME）" for row in rows))
        workbook.close()
        self.assertEqual(AuditEvent.objects.filter(event_type="REPORT_DOWNLOADED").count(), 1)

    def test_pdf_export_is_not_available(self):
        self.ingest()
        report = self.generate_and_publish()
        response = self.post_json(reverse("report-export-api", args=[report["reportId"]]), {"format": "PDF", "purpose": "不允许的PDF"})
        self.assertEqual(response.status_code, 422)
        self.assertEqual(response.json()["code"], "INVALID_EXPORT_FORMAT")

    def test_json_export_is_not_available(self):
        self.ingest()
        report = self.generate_and_publish()
        response = self.post_json(reverse("report-export-api", args=[report["reportId"]]), {"format": "JSON", "purpose": "不允许的完整数据"})
        self.assertEqual(response.status_code, 422)
        self.assertEqual(response.json()["code"], "INVALID_EXPORT_FORMAT")

    def test_expired_export_file_is_deleted_with_audit(self):
        self.ingest()
        report = self.generate_and_publish()
        response = self.post_json(reverse("report-export-api", args=[report["reportId"]]), {"format": "XLSX", "purpose": "短期测试文件"})
        job = ExportJob.objects.get(public_id=response.json()["data"]["exportId"])
        path = Path(job.file_path)
        job.expires_at = timezone.now() - timedelta(seconds=1)
        job.save(update_fields=["expires_at", "updated_at"])
        call_command("purge_expired_exports", verbosity=0)
        job.refresh_from_db()
        self.assertEqual(job.status, ExportJob.Status.DELETED)
        self.assertFalse(path.exists())
        self.assertTrue(AuditEvent.objects.filter(event_type="REPORT_EXPORT_DELETED", object_id=str(job.public_id)).exists())

    def test_out_of_scope_reporter_cannot_generate_or_download(self):
        self.ingest()
        report = self.generate_and_publish()
        export_response = self.post_json(reverse("report-export-api", args=[report["reportId"]]), {"format": "XLSX", "purpose": "授权企业内部使用"})
        export_id = export_response.json()["data"]["exportId"]
        self.client.force_login(self.outsider)
        generate = self.post_json(reverse("report-generate-api"), {"enterpriseId": str(self.enterprise.public_id), "periodType": "DAILY", "periodValue": "2026-07-19"})
        self.assertEqual(generate.status_code, 403)
        download = self.client.get(reverse("report-download-api", args=[export_id]))
        self.assertEqual(download.status_code, 403)

    def test_event_ingest_requires_shift_scope_and_action_token(self):
        self.client.force_login(self.monitor)
        no_token = self.client.post(reverse("report-event-upsert-api"), data=json.dumps({"event": event_payload(), "decision": decision_payload(), "action": {}}), content_type="application/json")
        self.assertEqual(no_token.status_code, 403)
        self.client.force_login(self.outsider)
        out_scope = self.post_json(reverse("report-event-upsert-api"), {"event": event_payload(), "decision": decision_payload(), "action": {}}, action_token=True)
        self.assertEqual(out_scope.status_code, 403)
        self.assertEqual(out_scope.json()["code"], "ENTERPRISE_SCOPE_DENIED")
        no_shift = get_user_model().objects.create_user(username="report-no-shift")
        AssistantProfile.objects.create(user=no_shift, display_name="无班次监控员", employee_code="RPT-NOSHIFT")
        EnterpriseGrant.objects.create(user=no_shift, enterprise=self.enterprise)
        assign_role(user=no_shift, role=RoleAssignment.Role.UNIT_USER, assigned_by=no_shift)
        self.client.force_login(no_shift)
        no_shift_response = self.post_json(reverse("report-event-upsert-api"), {"event": event_payload(), "decision": decision_payload(), "action": {}}, action_token=True)
        self.assertEqual(no_shift_response.status_code, 409)
        self.assertEqual(no_shift_response.json()["code"], "ACTIVE_SHIFT_REQUIRED")
