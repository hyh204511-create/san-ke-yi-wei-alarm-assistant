import hashlib
import json
import re
import secrets
import uuid
from collections import Counter, defaultdict
from copy import copy
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from django.conf import settings
from django.db import IntegrityError, transaction
from django.db.models import Max, Q
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill

from apps.governance.models import AuditEvent, DeviceRegistration, EnterpriseScope
from apps.governance.services import active_shift_for_user, enterprise_scope_for_user, enterprise_scope_ids_for_user

from .models import ExportJob, ReportSnapshot, ReportSourceBatch, ReportSourcePage, ReportTask, VehicleMonitorDailyRow
from .services import ReportingError, require_reporting_permission, require_scope


SOURCE_TYPES = {
    "ALARM_DISPOSAL_RATE",
    "ALARM_PROCESSING_RATE",
    "ALARM_CENTER",
    "VEHICLE_BASE_INFO",
    "TRACK_COMPLETENESS",
}
ALARM_SOURCES = ["ALARM_DISPOSAL_RATE", "ALARM_PROCESSING_RATE", "ALARM_CENTER"]
VEHICLE_SOURCES = ["VEHICLE_BASE_INFO", "TRACK_COMPLETENESS"]
SOURCE_SHEETS = {
    "ALARM_DISPOSAL_RATE": "处置率报表",
    "ALARM_PROCESSING_RATE": "处理率报表",
    "ALARM_CENTER": "报警查询报表",
}
ALARM_HEADERS = {
    "ALARM_DISPOSAL_RATE": ["序号", "机构", "类型", "车牌号", "车牌颜色", "所属地市", "正报总数", "申诉通过数", "判断中总数", "车辆报警总数", "处置率", "已处置报警数", "未处置报警数", "异常申诉数", "异常申诉通过数", "疑似事故报警", "夜间异动", "抽烟报警", "接打手持电话", "超员驾驶报警", "生理疲劳", "未系安全带", "超时驾驶", "驾驶员突发情况", "驾驶员身份识别报警", "押运员身份识别报警", "离线位移", "异地经营提醒", "电子围栏报警", "电子运单报警", "手部脱离方向盘报警", "设备故障报警", "超速驾驶", "分心驾驶"],
    "ALARM_PROCESSING_RATE": ["序号", "机构", "类型", "车牌号", "车牌颜色", "所属地市", "正报总数", "申诉通过总数", "判断中总数", "车辆报警总数", "处理率", "已处理报警数", "未处理报警数", "计入处理率报警", "未计入处理率报警", "异常申诉数", "异常申诉通过数", "疑似事故报警", "夜间异动", "抽烟报警", "接打手持电话", "超员驾驶报警", "生理疲劳", "未系安全带", "超时驾驶", "驾驶员突发情况", "驾驶员身份识别报警", "押运员身份识别报警", "离线位移", "异地经营提醒", "电子围栏报警", "电子运单报警", "手部脱离方向盘报警", "设备故障报警", "超速驾驶", "分心驾驶"],
    "ALARM_CENTER": ["序号", "状态", "报警ID", "类型", "车牌号", "终端版本", "车牌颜色", "驾驶员", "报警类型", "报警详情", "发生时间", "市州", "区县", "所属机构", "设备厂商", "设备型号", "定位速度(公里/时)", "仪表盘速度(公里/时)", "车型类型", "接收时间", "处置时间", "处理时间", "处置人", "处置方式", "处置内容", "复核人", "复核时间", "报警地址", "状态类型"],
}
SENSITIVE_KEY = re.compile(r"cookie|authorization|token|password|passwd|captcha|验证码|密钥", re.I)

# Real endpoint contracts intentionally remain disabled until an authorized
# operator walks each source and the captured shape is reviewed.
SOURCE_CONTRACTS = {
    source: {"enabled": False, "version": "UNVERIFIED", "route": "", "method": "", "path": ""}
    for source in SOURCE_TYPES
}


def canonical_json(value):
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)


def sha256_json(value):
    return hashlib.sha256(canonical_json(value).encode("utf-8")).hexdigest()


def reject_sensitive_keys(value, path="payload"):
    if isinstance(value, dict):
        for key, child in value.items():
            if SENSITIVE_KEY.search(str(key)):
                raise ReportingError(f"{path}包含禁止上传的认证字段", "SENSITIVE_FIELD_REJECTED", 422)
            reject_sensitive_keys(child, f"{path}.{key}")
    elif isinstance(value, list):
        for index, child in enumerate(value):
            reject_sensitive_keys(child, f"{path}[{index}]")


def value_shape(value):
    if isinstance(value, list):
        return [value_shape(value[0])] if value else []
    if isinstance(value, dict):
        return {key: value_shape(value[key]) for key in sorted(value)}
    return "value"


def rows_field_signature(rows):
    shapes = sorted({canonical_json(value_shape(row)) for row in rows[:20]})
    return sha256_json(shapes)


def required_sources(report_type):
    return ALARM_SOURCES if report_type in {
        ReportTask.ReportType.ALARM_DAILY,
        ReportTask.ReportType.ALARM_WEEKLY,
        ReportTask.ReportType.ALARM_MONTHLY,
    } else VEHICLE_SOURCES


def template_for(report_type):
    return "ALARM_THREE_SHEET_V1" if report_type.startswith("ALARM_") else "VEHICLE_MONITOR_DAILY_V1"


def period_for(report_type, start_value, end_value=None):
    try:
        start = date.fromisoformat(str(start_value))
        end = date.fromisoformat(str(end_value or start_value))
    except ValueError as exc:
        raise ReportingError("统计周期必须是有效ISO日期", "INVALID_REPORT_PERIOD", 422) from exc
    if report_type in {ReportTask.ReportType.ALARM_DAILY, ReportTask.ReportType.VEHICLE_MONITOR_DAILY} and end != start:
        raise ReportingError("日报周期必须为同一天", "INVALID_REPORT_PERIOD", 422)
    if report_type == ReportTask.ReportType.ALARM_WEEKLY and (start.weekday() != 0 or end != start + timedelta(days=6)):
        raise ReportingError("周报必须覆盖周一至周日", "INVALID_REPORT_PERIOD", 422)
    if report_type == ReportTask.ReportType.ALARM_MONTHLY:
        next_month = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
        if start.day != 1 or end != next_month - timedelta(days=1):
            raise ReportingError("月报必须覆盖完整自然月", "INVALID_REPORT_PERIOD", 422)
    return start, end


def query_conditions(report_type, start, end):
    common = {
        "vehicleStatuses": ["OPERATING", "OPERATING_NOT_ASSESSED"],
        "enterpriseSelection": "ALL_BY_VEHICLE_AFFILIATION",
        "periodStart": start.isoformat(),
        "periodEnd": end.isoformat(),
        "timezone": "Asia/Shanghai",
    }
    if report_type.startswith("ALARM_"):
        common.update({"dimension": "VEHICLE", "alarmTypeSelection": "ALL", "timeBasis": "SERVER_RECEIVED_AT"})
    else:
        common.update({"vehicleBaseHasNoDateFilter": True, "trackTargetDate": start.isoformat()})
    return common


def audit(actor, event_type, object_type, object_id, detail):
    AuditEvent.objects.create(
        actor=actor,
        event_type=event_type,
        object_type=object_type,
        object_id=str(object_id),
        role_snapshot=list(actor.assistant_roles.filter(is_active=True).values_list("role", flat=True)),
        enterprise_scope_snapshot=enterprise_scope_for_user(actor),
        detail=detail,
    )


@transaction.atomic
def create_report_task(*, actor, report_type, period_start, period_end=None, data_cutoff_at=None):
    require_reporting_permission(actor, "report.generate")
    if report_type not in ReportTask.ReportType.values:
        raise ReportingError("不支持的报表类型", "INVALID_REPORT_TYPE", 422)
    start, end = period_for(report_type, period_start, period_end)
    if report_type == ReportTask.ReportType.VEHICLE_MONITOR_DAILY and timezone.localdate() <= start:
        raise ReportingError("车辆日报只能在目标日次日生成", "REPORT_TOO_EARLY", 409)
    conditions = query_conditions(report_type, start, end)
    sources = required_sources(report_type)
    contracts = {source: SOURCE_CONTRACTS[source] for source in sources}
    query_spec = {"conditions": conditions, "contracts": contracts}
    cutoff = data_cutoff_at
    if isinstance(cutoff, str):
        cutoff = parse_datetime(cutoff.replace(" ", "T"))
        if not cutoff:
            raise ReportingError("数据截止时间格式无效", "INVALID_DATA_CUTOFF", 422)
        if timezone.is_naive(cutoff):
            cutoff = timezone.make_aware(cutoff, timezone.get_current_timezone())
    task = ReportTask.objects.create(
        report_type=report_type,
        period_start=start,
        period_end=end,
        target_date=start if report_type == ReportTask.ReportType.VEHICLE_MONITOR_DAILY else None,
        template_version=template_for(report_type),
        status=ReportTask.Status.WAITING_PLATFORM,
        query_spec=query_spec,
        required_source_types=sources,
        data_cutoff_at=cutoff or timezone.now(),
        requested_by=actor,
    )
    for source in sources:
        contract = contracts[source]
        source_conditions = {**conditions, "sourceType": source}
        ReportSourceBatch.objects.create(
            task=task,
            source_type=source,
            contract_version=contract["version"],
            query_hash=sha256_json(source_conditions),
            filters_summary=source_conditions,
        )
    audit(actor, "REPORT_TASK_CREATED", "REPORT_TASK", task.public_id, {"reportType": report_type, "periodStart": start.isoformat(), "periodEnd": end.isoformat()})
    return task


def task_contracts_verified(task):
    contracts = task.query_spec.get("contracts", {})
    return all(contracts.get(source, {}).get("enabled") is True for source in task.required_source_types)


@transaction.atomic
def claim_report_task(*, actor, task, device_id, duration_seconds=600):
    require_reporting_permission(actor, "report.collect", require_shift=True)
    task = ReportTask.objects.select_for_update().get(pk=task.pk)
    shift = active_shift_for_user(actor)
    if not task_contracts_verified(task):
        raise ReportingError("五来源真实接口契约尚未审核加入白名单", "REPORT_CONTRACT_UNVERIFIED", 409)
    device = DeviceRegistration.objects.filter(device_id=device_id, user=actor, is_active=True).first()
    if not device or device.session_status != "AUTHENTICATED":
        raise ReportingError("当前插件设备尚未确认省平台登录", "PLATFORM_SESSION_REQUIRED", 409)
    now = timezone.now()
    if task.lease_expires_at and task.lease_expires_at > now and task.device_id != device_id:
        raise ReportingError("该任务已被其他设备领取", "REPORT_TASK_LEASE_CONFLICT", 409)
    token = secrets.token_urlsafe(32)
    task.claimed_by = actor
    task.platform_account_ref = shift.platform_account_ref
    task.device_id = str(device_id)
    task.lease_token_hash = hashlib.sha256(token.encode()).hexdigest()
    task.lease_expires_at = now + timedelta(seconds=max(60, min(int(duration_seconds), 1800)))
    task.status = ReportTask.Status.FETCHING
    task.save(update_fields=["claimed_by", "platform_account_ref", "device_id", "lease_token_hash", "lease_expires_at", "status", "updated_at"])
    task._plain_lease_token = token
    audit(actor, "REPORT_TASK_CLAIMED", "REPORT_TASK", task.public_id, {"deviceId": device_id})
    return task


def verify_task_lease(task, actor, device_id, lease_token):
    if task.claimed_by_id != actor.pk or task.device_id != str(device_id or ""):
        raise ReportingError("报表任务设备或操作人不匹配", "REPORT_TASK_LEASE_DENIED", 403)
    expected = hashlib.sha256(str(lease_token or "").encode()).hexdigest()
    if not secrets.compare_digest(task.lease_token_hash or "", expected):
        raise ReportingError("报表任务租约无效", "REPORT_TASK_LEASE_DENIED", 403)
    if not task.lease_expires_at or task.lease_expires_at <= timezone.now():
        raise ReportingError("报表任务租约已过期", "REPORT_TASK_LEASE_EXPIRED", 409)


def validate_standard_rows(source_type, rows):
    if source_type not in SOURCE_TYPES or not isinstance(rows, list) or len(rows) > 5000:
        raise ReportingError("来源分页数据格式无效", "INVALID_REPORT_SOURCE_ROWS", 422)
    required = {
        "ALARM_DISPOSAL_RATE": {"enterpriseId", "enterpriseName", "values"},
        "ALARM_PROCESSING_RATE": {"enterpriseId", "enterpriseName", "values"},
        "ALARM_CENTER": {"enterpriseId", "enterpriseName", "values"},
        "VEHICLE_BASE_INFO": {"vehicleId", "plate", "enterpriseId", "enterpriseName", "vehicleStatus", "lastLocationTime"},
        "TRACK_COMPLETENESS": {"vehicleId", "plate", "enterpriseId", "enterpriseName", "totalMileage", "completeness"},
    }[source_type]
    for row in rows:
        if not isinstance(row, dict) or not required.issubset(row):
            raise ReportingError("标准行缺少必需字段", "REPORT_SOURCE_FIELD_MISSING", 422)
        if source_type in ALARM_SOURCES and not isinstance(row.get("values"), dict):
            raise ReportingError("报警来源values必须为对象", "INVALID_REPORT_SOURCE_ROWS", 422)


@transaction.atomic
def upload_source_page(*, actor, task, source_type, page_number, query_hash, field_signature, rows, device_id, lease_token):
    require_reporting_permission(actor, "report.collect", require_shift=True)
    reject_sensitive_keys(rows)
    validate_standard_rows(source_type, rows)
    task = ReportTask.objects.select_for_update().get(pk=task.pk)
    verify_task_lease(task, actor, device_id, lease_token)
    batch = ReportSourceBatch.objects.select_for_update().filter(task=task, source_type=source_type).first()
    if not batch or batch.query_hash != str(query_hash):
        raise ReportingError("请求条件哈希与冻结任务不一致", "REPORT_QUERY_HASH_MISMATCH", 409)
    page_number = int(page_number)
    if page_number < 1:
        raise ReportingError("页码必须从1开始", "INVALID_REPORT_PAGE", 422)
    computed_signature = rows_field_signature(rows)
    if computed_signature != str(field_signature):
        raise ReportingError("字段签名与标准行不一致", "REPORT_FIELD_SIGNATURE_MISMATCH", 409)
    page_hash = sha256_json(rows)
    existing = ReportSourcePage.objects.filter(task=task, source_type=source_type, page_number=page_number, query_hash=query_hash).first()
    if existing:
        if existing.page_hash != page_hash:
            raise ReportingError("同一分页幂等键提交了不同内容", "REPORT_PAGE_IDEMPOTENCY_CONFLICT", 409)
        return existing, False
    try:
        page = ReportSourcePage.objects.create(
            task=task, batch=batch, source_type=source_type, page_number=page_number,
            query_hash=query_hash, field_signature=field_signature, row_count=len(rows), page_hash=page_hash, rows=rows,
        )
    except IntegrityError as exc:
        raise ReportingError("来源分页并发冲突", "REPORT_PAGE_IDEMPOTENCY_CONFLICT", 409) from exc
    batch.status = ReportSourceBatch.Status.FETCHING
    batch.received_pages += 1
    batch.received_rows += len(rows)
    batch.save(update_fields=["status", "received_pages", "received_rows", "updated_at"])
    return page, True


@transaction.atomic
def complete_source_batch(*, actor, task, source_type, total_pages, total_rows, field_signature, device_id, lease_token):
    require_reporting_permission(actor, "report.collect", require_shift=True)
    task = ReportTask.objects.select_for_update().get(pk=task.pk)
    verify_task_lease(task, actor, device_id, lease_token)
    batch = ReportSourceBatch.objects.select_for_update().get(task=task, source_type=source_type)
    try:
        total_pages = int(total_pages)
        total_rows = int(total_rows)
    except (TypeError, ValueError) as exc:
        raise ReportingError("来源总页数和总行数必须是整数", "INVALID_REPORT_SOURCE_TOTAL", 422) from exc
    if total_pages < 1 or total_rows < 0:
        raise ReportingError("来源总页数或总行数无效", "INVALID_REPORT_SOURCE_TOTAL", 422)
    pages = list(batch.pages.order_by("page_number"))
    expected_pages = list(range(1, total_pages + 1))
    actual_pages = [page.page_number for page in pages]
    signatures = {page.field_signature for page in pages}
    problems = []
    if actual_pages != expected_pages:
        problems.append("分页不连续")
    if sum(page.row_count for page in pages) != total_rows:
        problems.append("分页行数之和与平台总行数不一致")
    if signatures != {str(field_signature)}:
        problems.append("分页字段签名不一致")
    batch.total_pages = total_pages
    batch.total_rows = total_rows
    batch.field_signature = str(field_signature)
    batch.status = ReportSourceBatch.Status.INVALID if problems else ReportSourceBatch.Status.COMPLETE
    batch.completed_at = timezone.now()
    batch.save(update_fields=["total_pages", "total_rows", "field_signature", "status", "completed_at", "updated_at"])
    if problems:
        task.status = ReportTask.Status.DATA_INCOMPLETE
        task.failure_code = "REPORT_SOURCE_INCOMPLETE"
        task.failure_reason = "；".join(problems)
        task.save(update_fields=["status", "failure_code", "failure_reason", "updated_at"])
    elif task.status == ReportTask.Status.DATA_INCOMPLETE and not task.source_batches.filter(status=ReportSourceBatch.Status.INVALID).exists():
        task.status = ReportTask.Status.FETCHING
        task.failure_code = ""
        task.failure_reason = ""
        task.save(update_fields=["status", "failure_code", "failure_reason", "updated_at"])
    return batch, problems


def source_rows(task, source_type, *, with_stats=False):
    rows = []
    seen = set()
    duplicate_rows = 0
    for page in ReportSourcePage.objects.filter(task=task, source_type=source_type).order_by("page_number"):
        for row in page.rows:
            row_hash = sha256_json(row)
            if row_hash in seen:
                duplicate_rows += 1
                continue
            seen.add(row_hash)
            rows.append(row)
    return (rows, duplicate_rows) if with_stats else rows


def resolve_row_enterprise(task, row):
    raw_id = str(row.get("enterpriseId") or "").strip()
    if not raw_id:
        return None, "ENTERPRISE_ID_MISSING"
    allowed = enterprise_scope_ids_for_user(task.requested_by)
    query = Q(code=raw_id)
    try:
        query |= Q(public_id=uuid.UUID(raw_id))
    except (ValueError, TypeError, AttributeError):
        pass
    matches = list(EnterpriseScope.objects.filter(query, is_active=True, pk__in=allowed))
    if len(matches) != 1:
        return None, "ENTERPRISE_UNKNOWN" if not matches else "ENTERPRISE_AMBIGUOUS"
    return matches[0], ""


def parse_platform_time(value):
    if value in (None, ""):
        return None
    parsed = parse_datetime(str(value).replace(" ", "T"))
    if not parsed:
        try:
            parsed = datetime.strptime(str(value), "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return "INVALID"
    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return timezone.localtime(parsed)


def decimal_value(value):
    try:
        return Decimal(str(value).replace("%", "").strip())
    except (InvalidOperation, AttributeError):
        return None


def next_snapshot_version(enterprise, report_type, start):
    return (ReportSnapshot.objects.filter(enterprise=enterprise, report_type=report_type, period_start=start).aggregate(Max("version"))["version__max"] or 0) + 1


def create_snapshot(task, enterprise, actor, metrics):
    period_type = {
        ReportTask.ReportType.ALARM_WEEKLY: ReportSnapshot.PeriodType.WEEKLY,
        ReportTask.ReportType.ALARM_MONTHLY: ReportSnapshot.PeriodType.MONTHLY,
    }.get(task.report_type, ReportSnapshot.PeriodType.DAILY)
    return ReportSnapshot.objects.create(
        task=task, report_type=task.report_type, template_version=task.template_version,
        enterprise=enterprise, period_type=period_type, period_start=task.period_start, period_end=task.period_end,
        version=next_snapshot_version(enterprise, task.report_type, task.period_start),
        status=ReportSnapshot.Status.DRAFT, review_status=ReportSnapshot.ReviewStatus.PENDING,
        metrics=metrics, parameters={"taskId": str(task.public_id)}, data_cutoff_at=task.data_cutoff_at,
        generated_by=actor,
    )


@transaction.atomic
def build_alarm_snapshots(task, actor):
    grouped = {source: defaultdict(list) for source in ALARM_SOURCES}
    anomalies = []
    enterprises = {}
    counts = {}
    for source in ALARM_SOURCES:
        rows, duplicate_rows = source_rows(task, source, with_stats=True)
        for index, row in enumerate(rows, 1):
            enterprise, reason = resolve_row_enterprise(task, row)
            if reason:
                anomalies.append({"sourceType": source, "rowNumber": index, "reason": reason})
                continue
            enterprises[enterprise.pk] = enterprise
            grouped[source][enterprise.pk].append(row)
        valid = sum(len(value) for value in grouped[source].values())
        anomaly_rows = len(rows) - valid
        counts[source] = {"total": len(rows) + duplicate_rows, "valid": valid, "duplicate": duplicate_rows, "anomaly": anomaly_rows}
        task.source_batches.filter(source_type=source).update(
            valid_rows=valid, duplicate_rows=duplicate_rows, anomaly_rows=anomaly_rows,
        )
        if valid != sum(len(value) for value in grouped[source].values()):
            raise ReportingError("企业拆分行数不守恒", "ENTERPRISE_ROW_CONSERVATION_FAILED", 409)
    snapshots = []
    for enterprise_id, enterprise in enterprises.items():
        metrics = {"sourceCounts": {source: len(grouped[source][enterprise_id]) for source in ALARM_SOURCES}}
        snapshots.append(create_snapshot(task, enterprise, actor, metrics))
    return snapshots, anomalies, counts


@transaction.atomic
def build_vehicle_rows(task, actor):
    VehicleMonitorDailyRow.objects.filter(task=task).delete()
    base_rows, base_duplicates = source_rows(task, "VEHICLE_BASE_INFO", with_stats=True)
    track_rows, track_duplicates = source_rows(task, "TRACK_COMPLETENESS", with_stats=True)
    base_by_id = defaultdict(list)
    track_by_id = defaultdict(list)
    for index, row in enumerate(base_rows):
        vehicle_id = str(row.get("vehicleId") or "").strip()
        base_by_id[vehicle_id or f"__missing_base_{index}"].append(row)
    for index, row in enumerate(track_rows):
        vehicle_id = str(row.get("vehicleId") or "").strip()
        track_by_id[vehicle_id or f"__missing_track_{index}"].append(row)
    target = task.target_date or task.period_start
    base_batch = task.source_batches.get(source_type="VEHICLE_BASE_INFO")
    track_batch = task.source_batches.get(source_type="TRACK_COMPLETENESS")
    created = []
    row_number = 0
    handled_track_ids = set()
    for _group_key, candidates in base_by_id.items():
        row_number += 1
        base = candidates[0]
        vehicle_id = str(base.get("vehicleId") or "").strip()
        tracks = track_by_id.get(vehicle_id, []) if vehicle_id else []
        if vehicle_id:
            handled_track_ids.add(vehicle_id)
        enterprise, enterprise_reason = resolve_row_enterprise(task, base)
        reasons = []
        if not vehicle_id:
            reasons.append("VEHICLE_ID_MISSING_PLATE_REVIEW_REQUIRED")
        if len(candidates) != 1:
            reasons.append("DUPLICATE_VEHICLE_ID")
        if len(tracks) > 1:
            reasons.append("TRACK_ONE_TO_MANY")
        track = tracks[0] if len(tracks) == 1 else None
        if enterprise_reason:
            reasons.append(enterprise_reason)
        if track and str(track.get("enterpriseId") or "") != str(base.get("enterpriseId") or ""):
            reasons.append("ENTERPRISE_CONFLICT")
        located = parse_platform_time(base.get("lastLocationTime"))
        online_status = ""
        trajectory_status = ""
        row_kind = VehicleMonitorDailyRow.RowKind.FORMAL
        filter_reason = ""
        if located == "INVALID":
            reasons.append("LAST_LOCATION_TIME_INVALID")
        elif located is None:
            if track:
                reasons.append("TRACK_PRESENT_WITHOUT_LAST_LOCATION")
            else:
                row_kind = VehicleMonitorDailyRow.RowKind.FILTERED
                filter_reason = "LAST_LOCATION_TIME_EMPTY"
                online_status = "过滤待确认"
        elif located.date() > target:
            reasons.append("LAST_LOCATION_AFTER_TARGET_DATE")
        elif located.date() == target:
            online_status = "上线"
            if not track:
                reasons.append("ONLINE_TRACK_MISSING")
        else:
            online_status = "未上线"
            if track:
                reasons.append("OFFLINE_TRACK_PRESENT")
        if track and not reasons:
            mileage = decimal_value(track.get("totalMileage"))
            completeness = decimal_value(track.get("completeness"))
            if mileage is None or mileage < 0 or completeness is None or completeness < 0 or completeness > 100:
                reasons.append("TRACK_VALUE_INVALID")
            elif online_status == "上线" and mileage == 0:
                trajectory_status = "在线无轨迹/未行驶"
            elif completeness == 100:
                trajectory_status = "正常"
            elif completeness < 100:
                trajectory_status = "轨迹不完整"
        elif online_status == "未上线" and not track:
            trajectory_status = "不适用"
        if reasons:
            row_kind = VehicleMonitorDailyRow.RowKind.ANOMALY
        data = {
            "vehicleId": vehicle_id,
            "plate": base.get("plate"),
            "enterpriseId": base.get("enterpriseId"),
            "enterpriseName": base.get("enterpriseName"),
            "vehicleStatus": base.get("vehicleStatus"),
            "targetDate": target.isoformat(),
            "lastLocationTime": base.get("lastLocationTime"),
            "totalMileage": track.get("totalMileage") if track else None,
            "completeness": track.get("completeness") if track else None,
            "onlineStatus": online_status,
            "trajectoryStatus": trajectory_status,
            "vehicleSourceBatchId": base_batch.pk,
            "trajectorySourceBatchId": track_batch.pk if track else None,
        }
        created.append(VehicleMonitorDailyRow.objects.create(
            task=task, enterprise=enterprise, row_number=row_number,
            vehicle_key_hash=hashlib.sha256((f"{vehicle_id}:duplicate:{row_number}" if len(candidates) > 1 else vehicle_id or f"plate:{base.get('plate')}:{row_number}").encode()).hexdigest(),
            row_kind=row_kind, online_status=online_status, trajectory_status=trajectory_status,
            filter_reason=filter_reason, anomaly_reason=";".join(reasons), critical=bool(reasons), data=data,
            vehicle_batch=base_batch, trajectory_batch=track_batch if track else None,
        ))
        for duplicate in candidates[1:]:
            row_number += 1
            duplicate_enterprise, duplicate_enterprise_reason = resolve_row_enterprise(task, duplicate)
            duplicate_reasons = ["DUPLICATE_VEHICLE_ID"]
            if duplicate_enterprise_reason:
                duplicate_reasons.append(duplicate_enterprise_reason)
            duplicate_data = {
                "vehicleId": vehicle_id, "plate": duplicate.get("plate"), "enterpriseId": duplicate.get("enterpriseId"),
                "enterpriseName": duplicate.get("enterpriseName"), "vehicleStatus": duplicate.get("vehicleStatus"),
                "targetDate": target.isoformat(), "lastLocationTime": duplicate.get("lastLocationTime"),
                "totalMileage": None, "completeness": None, "onlineStatus": "", "trajectoryStatus": "",
                "vehicleSourceBatchId": base_batch.pk, "trajectorySourceBatchId": None,
            }
            created.append(VehicleMonitorDailyRow.objects.create(
                task=task, enterprise=duplicate_enterprise, row_number=row_number,
                vehicle_key_hash=hashlib.sha256(f"{vehicle_id}:duplicate:{row_number}".encode()).hexdigest(),
                row_kind=VehicleMonitorDailyRow.RowKind.ANOMALY, anomaly_reason=";".join(duplicate_reasons),
                critical=True, data=duplicate_data, vehicle_batch=base_batch,
            ))
    for group_key, tracks in track_by_id.items():
        vehicle_id = str(tracks[0].get("vehicleId") or "").strip()
        if vehicle_id and vehicle_id in handled_track_ids:
            continue
        for track in tracks:
            row_number += 1
            enterprise, _reason = resolve_row_enterprise(task, track)
            data = {**track, "targetDate": target.isoformat(), "onlineStatus": "", "trajectoryStatus": ""}
            created.append(VehicleMonitorDailyRow.objects.create(
                task=task, enterprise=enterprise, row_number=row_number,
                vehicle_key_hash=hashlib.sha256((vehicle_id or f"unmatched:{row_number}").encode()).hexdigest(),
                row_kind=VehicleMonitorDailyRow.RowKind.ANOMALY, anomaly_reason="TRACK_WITHOUT_VEHICLE_BASE",
                critical=True, data=data, vehicle_batch=base_batch, trajectory_batch=track_batch,
            ))
    snapshots = []
    enterprise_ids = sorted({row.enterprise_id for row in created if row.enterprise_id})
    for enterprise_id in enterprise_ids:
        enterprise = EnterpriseScope.objects.get(pk=enterprise_id)
        counts = Counter(row.row_kind for row in created if row.enterprise_id == enterprise_id)
        snapshot = create_snapshot(task, enterprise, actor, {"rowCounts": dict(counts)})
        VehicleMonitorDailyRow.objects.filter(task=task, enterprise=enterprise).update(snapshot=snapshot)
        snapshots.append(snapshot)
    base_anomalies = sum(1 for row in created if row.vehicle_batch_id == base_batch.pk and row.critical)
    track_anomalies = sum(1 for row in created if row.trajectory_batch_id == track_batch.pk and row.critical)
    ReportSourceBatch.objects.filter(pk=base_batch.pk).update(
        valid_rows=max(0, len(base_rows) - base_anomalies), duplicate_rows=base_duplicates, anomaly_rows=base_anomalies,
    )
    ReportSourceBatch.objects.filter(pk=track_batch.pk).update(
        valid_rows=max(0, len(track_rows) - track_anomalies), duplicate_rows=track_duplicates, anomaly_rows=track_anomalies,
    )
    return snapshots, created


@transaction.atomic
def finalize_report_task(*, actor, task):
    require_reporting_permission(actor, "report.collect", require_shift=True)
    task = ReportTask.objects.select_for_update().get(pk=task.pk)
    if task.status != ReportTask.Status.FETCHING:
        raise ReportingError("只有取数中的任务可以执行最终校验", "INVALID_REPORT_TASK_STATUS", 409)
    batches = list(task.source_batches.all())
    if len(batches) != len(task.required_source_types) or any(batch.status != ReportSourceBatch.Status.COMPLETE for batch in batches):
        task.status = ReportTask.Status.DATA_INCOMPLETE
        task.failure_code = "REPORT_SOURCES_INCOMPLETE"
        task.failure_reason = "五来源任务存在未完成或无效来源"
        task.save(update_fields=["status", "failure_code", "failure_reason", "updated_at"])
        raise ReportingError(task.failure_reason, task.failure_code, 409)
    task.status = ReportTask.Status.VALIDATING
    task.save(update_fields=["status", "updated_at"])
    task.snapshots.filter(status=ReportSnapshot.Status.DRAFT).delete()
    if task.report_type.startswith("ALARM_"):
        snapshots, anomalies, counts = build_alarm_snapshots(task, actor)
        critical = len(anomalies)
        summary = {"sourceCounts": counts, "anomalies": anomalies, "snapshotCount": len(snapshots)}
    else:
        snapshots, rows = build_vehicle_rows(task, actor)
        critical = sum(1 for row in rows if row.critical)
        summary = {
            "rowCounts": dict(Counter(row.row_kind for row in rows)),
            "criticalIssueCount": critical,
            "snapshotCount": len(snapshots),
        }
    if not snapshots:
        critical += 1
        summary["taskIssue"] = "NO_REPORTABLE_ENTERPRISE_DATA"
    task.status = ReportTask.Status.REVIEW_REQUIRED
    task.critical_issue_count = critical
    task.validation_summary = summary
    task.failure_code = ""
    task.failure_reason = ""
    task.save(update_fields=["status", "critical_issue_count", "validation_summary", "failure_code", "failure_reason", "updated_at"])
    audit(actor, "REPORT_TASK_VALIDATED", "REPORT_TASK", task.public_id, {"criticalIssueCount": critical, "snapshotCount": len(snapshots)})
    return task


@transaction.atomic
def review_report_task(*, actor, task, approve, note=""):
    require_reporting_permission(actor, "report.publish")
    task = ReportTask.objects.select_for_update().get(pk=task.pk)
    if task.status != ReportTask.Status.REVIEW_REQUIRED:
        raise ReportingError("只有待审核任务可以审核", "INVALID_REPORT_TASK_STATUS", 409)
    allowed = enterprise_scope_ids_for_user(actor)
    if not set(task.snapshots.values_list("enterprise_id", flat=True)).issubset(allowed):
        raise ReportingError("无权审核该任务的全部企业报表", "ENTERPRISE_SCOPE_DENIED", 403)
    if approve and task.critical_issue_count:
        raise ReportingError("任务仍有关键异常，不能正式发布", "REPORT_CRITICAL_ISSUES_UNRESOLVED", 409)
    now = timezone.now()
    task.status = ReportTask.Status.APPROVED if approve else ReportTask.Status.REJECTED
    task.reviewed_by = actor
    task.reviewed_at = now
    task.review_note = str(note or "")[:1000]
    task.save(update_fields=["status", "reviewed_by", "reviewed_at", "review_note", "updated_at"])
    for snapshot in task.snapshots.select_for_update():
        snapshot.review_status = ReportSnapshot.ReviewStatus.APPROVED if approve else ReportSnapshot.ReviewStatus.REJECTED
        if approve:
            ReportSnapshot.objects.filter(
                enterprise=snapshot.enterprise, report_type=snapshot.report_type, period_type=snapshot.period_type,
                period_start=snapshot.period_start, status=ReportSnapshot.Status.PUBLISHED,
            ).exclude(pk=snapshot.pk).update(status=ReportSnapshot.Status.RETIRED)
            snapshot.status = ReportSnapshot.Status.PUBLISHED
            snapshot.published_by = actor
            snapshot.published_at = now
        snapshot.save(update_fields=["review_status", "status", "published_by", "published_at", "updated_at"])
    audit(actor, "REPORT_TASK_APPROVED" if approve else "REPORT_TASK_REJECTED", "REPORT_TASK", task.public_id, {"note": task.review_note})
    return task


def template_path(version):
    return Path(__file__).resolve().parent / "report_templates" / f"{version}.xlsx"


def copy_row_style(source, target):
    if source.has_style:
        target._style = copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy(source.alignment)


def alarm_rows_for_snapshot(snapshot, source):
    result = []
    for row in source_rows(snapshot.task, source):
        enterprise, reason = resolve_row_enterprise(snapshot.task, row)
        if not reason and enterprise.pk == snapshot.enterprise_id:
            result.append(row)
    return result


def build_alarm_xlsx(snapshot):
    path = template_path("ALARM_THREE_SHEET_V1")
    workbook = load_workbook(path) if path.exists() else Workbook()
    if not path.exists():
        workbook.remove(workbook.active)
        for source in ALARM_SOURCES:
            sheet = workbook.create_sheet(SOURCE_SHEETS[source])
            sheet.append(["企业报警报表"])
            sheet.append(ALARM_HEADERS[source])
    title = f"{snapshot.enterprise.name}{snapshot.period_start}至{snapshot.period_end}"
    for source in ALARM_SOURCES:
        sheet = workbook[SOURCE_SHEETS[source]]
        sheet.cell(1, 1, title)
        style_row = [copy(sheet.cell(3, column)._style) for column in range(1, len(ALARM_HEADERS[source]) + 1)]
        for row in sheet.iter_rows(min_row=3):
            for cell in row:
                cell.value = None
        headers = ALARM_HEADERS[source]
        for row_index, row in enumerate(alarm_rows_for_snapshot(snapshot, source), 3):
            values = row.get("values", {})
            for column, header in enumerate(headers, 1):
                cell = sheet.cell(row_index, column, values.get(header))
                cell._style = copy(style_row[column - 1])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


AUDIT_HEADERS = ["车辆ID", "车牌号", "企业ID", "企业名称", "车辆状态", "目标日", "最后定位时间", "上线状态", "总里程", "轨迹完整率", "轨迹状态", "车辆来源批次", "轨迹来源批次", "过滤原因", "异常原因"]


def append_vehicle_detail(sheet, rows):
    sheet.append(AUDIT_HEADERS)
    for row in rows:
        data = row.data
        sheet.append([
            data.get("vehicleId"), data.get("plate"), data.get("enterpriseId"), data.get("enterpriseName"), data.get("vehicleStatus"),
            data.get("targetDate"), data.get("lastLocationTime"), data.get("onlineStatus"), data.get("totalMileage"), data.get("completeness"),
            data.get("trajectoryStatus"), data.get("vehicleSourceBatchId"), data.get("trajectorySourceBatchId"), row.filter_reason, row.anomaly_reason,
        ])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"


def build_vehicle_xlsx(snapshot):
    path = template_path("VEHICLE_MONITOR_DAILY_V1")
    workbook = load_workbook(path) if path.exists() else Workbook()
    sheet = workbook.active
    sheet.title = (snapshot.enterprise.name[:20] + "车辆动态监控日报表")[:31]
    sheet.cell(1, 1, f"{snapshot.enterprise.name}车辆动态监控日报表")
    style_row = [copy(sheet.cell(6, column)._style) for column in range(1, 13)]
    row_height = sheet.row_dimensions[6].height
    for row in sheet.iter_rows(min_row=6):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None
    rows = list(VehicleMonitorDailyRow.objects.filter(snapshot=snapshot).order_by("row_number"))
    formal = [row for row in rows if row.row_kind == VehicleMonitorDailyRow.RowKind.FORMAL]
    for index, row in enumerate(formal, 1):
        excel_row = index + 5
        data = row.data
        sheet.cell(excel_row, 1, index)
        sheet.cell(excel_row, 2, data.get("plate"))
        sheet.cell(excel_row, 6, row.online_status)
        sheet.cell(excel_row, 7, data.get("lastLocationTime"))
        sheet.cell(excel_row, 8, data.get("completeness"))
        for column in range(1, 13):
            sheet.cell(excel_row, column)._style = copy(style_row[column - 1])
        sheet.row_dimensions[excel_row].height = row_height
    for name in ["明细审计", "过滤清单", "异常清单"]:
        if name in workbook.sheetnames:
            del workbook[name]
    append_vehicle_detail(workbook.create_sheet("明细审计"), formal)
    append_vehicle_detail(workbook.create_sheet("过滤清单"), [row for row in rows if row.row_kind == VehicleMonitorDailyRow.RowKind.FILTERED])
    append_vehicle_detail(workbook.create_sheet("异常清单"), [row for row in rows if row.row_kind == VehicleMonitorDailyRow.RowKind.ANOMALY])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def snapshot_xlsx(snapshot):
    if snapshot.status != ReportSnapshot.Status.PUBLISHED or snapshot.review_status != ReportSnapshot.ReviewStatus.APPROVED:
        raise ReportingError("只有已审核报表可以导出", "REPORT_NOT_APPROVED", 409)
    return build_alarm_xlsx(snapshot) if snapshot.report_type.startswith("ALARM_") else build_vehicle_xlsx(snapshot)


def safe_file_part(value):
    return re.sub(r"[\\/:*?\"<>|\s]+", "-", str(value)).strip("-")[:80] or "report"


def write_export(content, file_name):
    export_dir = Path(settings.REPORT_EXPORT_DIR).resolve()
    export_dir.mkdir(parents=True, exist_ok=True)
    stem = Path(file_name).stem
    suffix = Path(file_name).suffix
    unique_name = f"{stem}-{uuid.uuid4().hex[:12]}{suffix}"
    path = (export_dir / unique_name).resolve()
    if export_dir not in path.parents:
        raise ReportingError("导出路径无效", "INVALID_EXPORT_PATH", 500)
    path.write_bytes(content)
    return path


@transaction.atomic
def create_snapshot_export(*, actor, snapshot, purpose):
    require_reporting_permission(actor, "export.masked")
    require_scope(actor, snapshot.enterprise)
    content = snapshot_xlsx(snapshot)
    file_name = f"{safe_file_part(snapshot.enterprise.name)}-{snapshot.report_type}-{snapshot.period_start}-v{snapshot.version}.xlsx"
    path = write_export(content, file_name)
    file_name = path.name
    job = ExportJob.objects.create(
        report_snapshot=snapshot, report_task=None, format=ExportJob.Format.XLSX, purpose=str(purpose or "报表下载")[:500],
        file_name=file_name, file_path=str(path), file_sha256=hashlib.sha256(content).hexdigest(), file_size=len(content),
        created_by=actor, expires_at=timezone.now() + timedelta(days=settings.REPORT_EXPORT_RETENTION_DAYS),
    )
    audit(actor, "REPORT_EXPORTED", "EXPORT_JOB", job.public_id, {"snapshotId": str(snapshot.public_id), "fileHash": job.file_sha256})
    return job


def global_anomaly_xlsx(task):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "全局异常清单"
    sheet.append(["来源", "行号", "异常原因"])
    for item in task.validation_summary.get("anomalies", []):
        sheet.append([item.get("sourceType"), item.get("rowNumber"), item.get("reason")])
    for row in VehicleMonitorDailyRow.objects.filter(task=task, enterprise__isnull=True).order_by("row_number"):
        sheet.append(["VEHICLE_MONITOR_DAILY", row.row_number, row.anomaly_reason])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@transaction.atomic
def create_task_bundle_export(*, actor, task, purpose):
    require_reporting_permission(actor, "export.masked")
    allowed = enterprise_scope_ids_for_user(actor)
    if not set(task.snapshots.values_list("enterprise_id", flat=True)).issubset(allowed):
        raise ReportingError("无权导出该任务的全部企业报表", "ENTERPRISE_SCOPE_DENIED", 403)
    if task.status != ReportTask.Status.APPROVED:
        raise ReportingError("只有已审核任务可以批量导出", "REPORT_NOT_APPROVED", 409)
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        for snapshot in task.snapshots.select_related("enterprise").order_by("enterprise__code"):
            name = f"{safe_file_part(snapshot.enterprise.name)}-{snapshot.report_type}-{snapshot.period_start}.xlsx"
            archive.writestr(name, snapshot_xlsx(snapshot))
        archive.writestr("_全局异常清单.xlsx", global_anomaly_xlsx(task))
        archive.writestr("manifest.json", canonical_json({
            "taskId": str(task.public_id), "reportType": task.report_type, "templateVersion": task.template_version,
            "periodStart": task.period_start.isoformat(), "periodEnd": task.period_end.isoformat(),
        }))
    content = output.getvalue()
    file_name = f"{task.report_type}-{task.period_start}-bundle.zip"
    path = write_export(content, file_name)
    file_name = path.name
    job = ExportJob.objects.create(
        report_snapshot=None, report_task=task, format=ExportJob.Format.ZIP, purpose=str(purpose or "批量下载")[:500],
        file_name=file_name, file_path=str(path), file_sha256=hashlib.sha256(content).hexdigest(), file_size=len(content),
        created_by=actor, expires_at=timezone.now() + timedelta(days=settings.REPORT_EXPORT_RETENTION_DAYS),
    )
    audit(actor, "REPORT_BUNDLE_EXPORTED", "EXPORT_JOB", job.public_id, {"taskId": str(task.public_id), "fileHash": job.file_sha256})
    return job
